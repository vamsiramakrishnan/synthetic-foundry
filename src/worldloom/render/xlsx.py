"""The XLSX renderer.

First of the renderers, because the workbook is a source artifact rather than a
projection of one, and because it is the only format where numerical coherence is
externally checkable: open the file, and the sheet recomputes its own totals.

**Formulas remain formulas.** A total is written as ``=SUM(C4:C6)``, not as the
number that sum happens to equal. A variance is ``=D4-C4``. A margin is
``=F4/E4``. The literal value is still in the IR and is what a formula-free format
emits, but nothing here pastes a computed value into a cell that should compute.

That is the difference between a workbook and a screenshot of one. If a fact
changes and a total does not, the sheet shows it.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

from ..models import ArtifactIR, FormulaKind, Table
from . import Rendered, RenderError, ooxml, slug_for

if TYPE_CHECKING:  # pragma: no cover
    from ..world import World

#: Artifact types this renderer handles. Everything else belongs to another
#: format. A mutable set with a registration function, not a frozenset: which
#: types are workbooks is domain vocabulary, and a vertical's source artifact
#: (banking's capital return) registers from its own module rather than being
#: named here — the same seam `documents.register_artifact_types` provides.
HANDLES = {"finance_workbook"}


def register(*artifact_types: str) -> None:
    """Claim *artifact_types* for the XLSX renderer.

    Called at domain-module import, like every registration seam, so a corpus
    renders the same set of files in every process.
    """
    HANDLES.update(artifact_types)

_HEADER_ROW = 3
"""Row 1 is the title, row 2 the subtitle, row 3 the column headers."""


def _require_openpyxl():  # type: ignore[no-untyped-def]
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on install extras
        raise RenderError(
            "the xlsx format needs openpyxl. Install it with: pip install 'worldloom[xlsx]'"
        ) from exc
    return openpyxl


def _column_letter(index: int) -> str:
    """1-based column index to its spreadsheet letter."""
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


class _Layout:
    """Where every row and column of every table landed on the sheet.

    Built once per workbook rather than searched per cell. A store-level sheet has
    thousands of rows and three formula columns; resolving each operand by
    scanning the row list is quadratic and turns a one-second render into a
    one-minute one.
    """

    __slots__ = ("rows", "columns", "first_data_row", "sheet")

    def __init__(self, tables, sheet_of: dict[str, str], rows_of: dict[str, int]) -> None:
        self.rows = {
            table.key: {row.key: index for index, row in enumerate(table.rows)}
            for table in tables
        }
        self.columns = {
            table.key: {column.key: index for index, column in enumerate(table.columns)}
            for table in tables
        }
        self.first_data_row = rows_of
        self.sheet = sheet_of

    def cell(self, table_key: str, row_key: str, column_key: str) -> tuple[int, str] | None:
        """``(row index, A1 address)``, or ``None`` if either key is unknown."""
        row_index = self.rows.get(table_key, {}).get(row_key)
        column_index = self.columns.get(table_key, {}).get(column_key)
        if row_index is None or column_index is None:
            return None
        # +2 for the label column that precedes the data columns.
        return row_index, f"{_column_letter(column_index + 2)}{self.first_data_row[table_key] + row_index}"


def _operand(table_key: str, operand: str, column_key: str) -> tuple[str, str, str]:
    """Split a SUM operand into ``(table, row, column)``.

    A bare operand names a row in the current table and the column being
    computed. A ``table:row:column`` operand names a cell anywhere in the
    workbook, which is how the reconciliation sheet sums the P&L's own rows
    instead of restating them.
    """
    parts = operand.split(":")
    if len(parts) == 3:
        return parts[0], parts[1], parts[2]
    return table_key, operand, column_key


def _formula(
    ir: ArtifactIR,
    table: Table,
    row_key: str,
    column_key: str,
    cell,  # type: ignore[no-untyped-def]
    *,
    layout: _Layout,
) -> str | None:
    """Translate a declared computation into Excel syntax.

    The IR says *what* is computed; this decides how to spell it. Nothing here
    invents a computation the IR did not declare.
    """
    if cell.formula is None:
        return None

    def here(target_column: str, target_row: str | None = None) -> str | None:
        found = layout.cell(table.key, target_row or row_key, target_column)
        return found[1] if found else None

    if cell.formula is FormulaKind.SUM:
        resolved = []
        for operand in cell.operands:
            target_table, target_row, target_column = _operand(table.key, operand, column_key)
            found = layout.cell(target_table, target_row, target_column)
            if found is not None:
                resolved.append((target_table, target_column, found[0], found[1]))
        if not resolved:
            return None

        def qualify(entry) -> str:  # type: ignore[no-untyped-def]
            target_table, _, _, address = entry
            if target_table == table.key:
                return address
            return f"'{layout.sheet[target_table]}'!{address}"

        # A range only when the operands really are consecutive rows of one
        # column. Collapsing a scattered set into `first:last` would silently
        # include the rows in between — with subtotal rows on the sheet, that is
        # every category counted twice.
        contiguous = (
            len(resolved) > 1
            and len({(t, c) for t, c, _, _ in resolved}) == 1
            and all(b[2] == a[2] + 1 for a, b in zip(resolved, resolved[1:]))
        )
        if contiguous:
            target_table = resolved[0][0]
            first, last = resolved[0][3], resolved[-1][3]
            if target_table == table.key:
                return f"=SUM({first}:{last})"
            # The sheet is named once for the range, not once per endpoint: Excel
            # rejects `'Sheet'!C4:'Sheet'!C6`.
            return f"=SUM('{layout.sheet[target_table]}'!{first}:{last})"
        return f"=SUM({','.join(qualify(entry) for entry in resolved)})"

    if cell.formula is FormulaKind.DIFFERENCE and len(cell.operands) == 2:
        left, right = here(cell.operands[0]), here(cell.operands[1])
        return f"={left}-{right}" if left and right else None

    if cell.formula is FormulaKind.RATIO_PCT and len(cell.operands) == 2:
        numerator, denominator = here(cell.operands[0]), here(cell.operands[1])
        if not (numerator and denominator):
            return None
        # Guarded, because a zero denominator in a spreadsheet is a #DIV/0! a
        # reader has to interpret rather than a number they can read.
        return f"=IF({denominator}=0,0,{numerator}/{denominator})"

    if cell.formula is FormulaKind.REFERENCE and cell.operands:
        target_table, target_row, target_column = _operand(table.key, cell.operands[0], column_key)
        found = layout.cell(target_table, target_row, target_column)
        if found is None or target_table not in layout.sheet:
            return None
        return f"='{layout.sheet[target_table]}'!{found[1]}"

    return None


def render(ir: ArtifactIR, detail=()) -> bytes:  # type: ignore[no-untyped-def]
    """Render one workbook IR to XLSX bytes.

    *detail* is the sequence of ``detail.DetailTable`` rows bound to this
    artifact — each becomes a real sheet after the IR's own sections, which is
    what turns a workbook of fifteen load-bearing rows into something shaped
    like a system export. Passed in rather than read from a world because this
    function renders one IR in isolation (the determinism tests call it that
    way), and defaulted empty so every workbook without a detail recipe keeps
    its exact bytes.
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import quote_sheetname

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    sections = [s for s in ir.sections if s.table is not None]
    sheet_of = {s.table.key: s.table.title[:31] for s in sections}
    rows_of = {s.table.key: _HEADER_ROW + 1 for s in sections}
    layout = _Layout([s.table for s in sections], sheet_of, rows_of)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")

    for section in sections:
        table = section.table
        sheet = workbook.create_sheet(title=sheet_of[table.key])
        sheet.sheet_state = "hidden" if section.hidden else "visible"

        sheet.cell(row=1, column=1, value=ir.title).font = Font(bold=True, size=13)
        if ir.subtitle:
            sheet.cell(row=2, column=1, value=ir.subtitle).font = Font(italic=True, color="666666")

        sheet.cell(row=_HEADER_ROW, column=1, value=table.title).font = bold
        sheet.cell(row=_HEADER_ROW, column=1).fill = header_fill
        for index, column in enumerate(table.columns, start=2):
            header = sheet.cell(row=_HEADER_ROW, column=index, value=column.label)
            header.font = bold
            header.fill = header_fill
            header.alignment = Alignment(horizontal="right")

        for row_offset, row in enumerate(table.rows):
            excel_row = rows_of[table.key] + row_offset
            label = sheet.cell(row=excel_row, column=1, value=row.label)
            if row.emphasis:
                label.font = bold

            for column_index, column in enumerate(table.columns, start=2):
                cell = row.cells.get(column.key)
                target = sheet.cell(row=excel_row, column=column_index)
                if cell is None:
                    continue

                formula = _formula(ir, table, row.key, column.key, cell, layout=layout)

                if formula is not None:
                    target.value = formula
                else:
                    target.value = cell.value

                if column.number_format:
                    # A percentage fact is stored as 24.94, and Excel's percent
                    # format multiplies by 100 — so scale to a fraction rather
                    # than showing 2494%.
                    if column.number_format.endswith("%") and isinstance(cell.value, (int, float)):
                        if formula is not None:
                            pass  # a ratio formula already yields a fraction
                        else:
                            target.value = cell.value / 100
                    target.number_format = column.number_format
                if row.emphasis:
                    target.font = bold

        widths = [max(12, len(table.title))] + [max(12, len(c.label) + 2) for c in table.columns]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[_column_letter(index)].width = width

        if table.note:
            note_row = rows_of[table.key] + len(table.rows) + 1
            sheet.cell(row=note_row, column=1, value=table.note).font = Font(
                italic=True, color="666666"
            )

        sheet.freeze_panes = sheet.cell(row=rows_of[table.key], column=2)

    for table in detail:
        _detail_sheet(workbook, ir, table, bold=bold, header_fill=header_fill,
                      Alignment=Alignment, Font=Font)

    # Named ranges, so a consumer can address the numbers that matter without
    # depending on where they happen to sit.
    pnl = next((t for t in ir.tables() if t.key == "pnl"), None)
    if pnl is not None:
        group = next((row for row in reversed(pnl.rows) if row.emphasis), None)
        if group is not None:
            sheet_name = sheet_of["pnl"]
            for column_key, name in (
                ("revenue_actual", "GroupRevenueActual"),
                ("revenue_budget", "GroupRevenueBudget"),
                ("revenue_variance", "GroupRevenueVariance"),
                ("gp_actual", "GroupGrossProfitActual"),
                ("gp_variance", "GroupGrossProfitVariance"),
            ):
                found = layout.cell("pnl", group.key, column_key)
                address = found[1] if found else None
                if address:
                    workbook.defined_names[name] = DefinedName(
                        name, attr_text=f"{quote_sheetname(sheet_name)}!${address[0]}${address[1:]}"
                    )

    _add_charts(workbook, ir, layout, sheet_of)

    workbook.properties.title = ir.title
    workbook.properties.subject = ir.subtitle or ""
    workbook.properties.creator = "Worldloom"
    workbook.properties.description = ir.metadata.get(
        "note", "Synthetic corpus generated by Worldloom."
    )
    workbook.properties.keywords = "synthetic,worldloom,seed=" + ir.metadata.get("worldloom_seed", "")

    # Document timestamps come from the world, never from the clock. openpyxl
    # defaults these to `datetime.now()`, which would put the moment of rendering
    # inside the file and break byte-identical regeneration.
    stamp = ir.metadata.get("worldloom_created")
    if stamp:
        moment = datetime.fromisoformat(stamp).replace(tzinfo=None)
        workbook.properties.created = moment
        workbook.properties.modified = moment

    buffer = BytesIO()
    workbook.save(buffer)
    return ooxml.normalise(buffer.getvalue(), created=stamp)


def _detail_sheet(workbook, ir, table, *, bold, header_fill, Alignment, Font):  # type: ignore[no-untyped-def]
    """One detail table as a sheet of literal rows plus a computed total.

    Values are literals — a thousand generated lines are the data, not a
    projection of other cells — but the **total row is a formula**, exactly as
    every other sheet keeps its totals: the workbook's claim is that a reader
    can recompute it, and for a fact-backed column the sum the formula shows
    is the ledger's own figure, because the rows were allocated from it.
    """
    sheet = workbook.create_sheet(title=table.title[:31])
    sheet.cell(row=1, column=1, value=ir.title).font = Font(bold=True, size=13)
    sheet.cell(row=2, column=1, value=f"{table.title} · {len(table.rows):,} lines").font = Font(
        italic=True, color="666666"
    )

    for index, column in enumerate(table.columns, start=1):
        header = sheet.cell(row=_HEADER_ROW, column=index, value=column.label)
        header.font = bold
        header.fill = header_fill
        if column.number_format:
            header.alignment = Alignment(horizontal="right")

    first = _HEADER_ROW + 1
    for offset, row in enumerate(table.rows):
        for index, column in enumerate(table.columns, start=1):
            cell = sheet.cell(row=first + offset, column=index, value=row.get(column.name))
            if column.number_format:
                cell.number_format = column.number_format

    total_row = first + len(table.rows)
    sheet.cell(row=total_row, column=1, value="Total").font = bold
    for index, column in enumerate(table.columns, start=1):
        if not column.fact_id:
            continue
        letter = _column_letter(index)
        cell = sheet.cell(row=total_row, column=index)
        cell.value = f"=SUM({letter}{first}:{letter}{total_row - 1})"
        cell.font = bold
        if column.number_format:
            cell.number_format = column.number_format

    if table.note:
        sheet.cell(row=total_row + 2, column=1, value=table.note).font = Font(
            italic=True, color="666666"
        )

    widths = [max(12, len(c.label) + 2) for c in table.columns]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=first, column=1)


def render_all(world: World) -> list[Rendered]:
    """Render every workbook in *world*, with its detail sheets attached."""
    by_intent: dict[str, list] = {}
    for table in world.detail_tables:
        if table.artifact_id:
            by_intent.setdefault(table.artifact_id, []).append(table)

    out: list[Rendered] = []
    for ir in world.artifact_irs:
        intent = world.artifact_intents.by_id(ir.intent_id)
        if intent.artifact_type not in HANDLES:
            continue
        out.append(
            Rendered(
                artifact_id=ir.id,
                path=f"artifacts/{ir.id.lower()}-{slug_for(intent.artifact_type)}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                payload=render(ir, detail=by_intent.get(ir.intent_id, ())),
            )
        )
    return out


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------

#: Where the chart source block lives. Hidden, because it is machinery rather
#: than content — but present, and made of live cross-sheet formulas rather than
#: pasted values, so the chart still moves when a figure does.
_CHART_DATA = "Chart Data"

_KIND = {"column": ("BarChart", "col"), "bar": ("BarChart", "bar"), "line": ("LineChart", None), "pie": ("PieChart", None)}


def _chart_block(sheet, chart, table, layout, *, at_row: int):  # type: ignore[no-untyped-def]
    """Write one chart's source block, and return its extent.

    The block is not a copy of the data. Every value cell is a formula pointing at
    the cell it charts, so a reader who changes a figure sees the chart move, and
    a chart can never quietly disagree with the table it claims to plot. That is
    the same reason the workbook keeps formulas everywhere else.

    A block is needed at all because a chart's rows are frequently not contiguous
    — the category chart plots thirty-four categories with unit subtotals sitting
    between them — and a spreadsheet range is a rectangle.
    """
    source = layout.sheet[chart.table]

    if chart.by_row:
        categories, series = list(chart.series), list(chart.rows)
    else:
        categories, series = list(chart.rows), list(chart.series)

    def label(key: str, is_row: bool) -> str:
        if is_row:
            row = next((r for r in table.rows if r.key == key), None)
            return row.label if row else key
        column = table.column(key)
        return column.label if column else key

    sheet.cell(row=at_row, column=1, value=chart.title)
    for index, key in enumerate(series, start=2):
        sheet.cell(row=at_row, column=index, value=label(key, chart.by_row))

    for offset, category in enumerate(categories, start=1):
        sheet.cell(row=at_row + offset, column=1, value=label(category, not chart.by_row))
        for index, member in enumerate(series, start=2):
            # `by_row` decides which of the pair names a row and which names a
            # column. Resolved explicitly, because getting it backwards produces a
            # block of empty cells and a chart that draws nothing while looking
            # entirely well-formed.
            row_key, column_key = (member, category) if chart.by_row else (category, member)
            found = layout.cell(chart.table, row_key, column_key)
            if found is not None:
                sheet.cell(row=at_row + offset, column=index).value = f"='{source}'!{found[1]}"

    return at_row + len(categories), len(series)


def _add_charts(workbook, ir, layout, sheet_of):  # type: ignore[no-untyped-def]
    """Draw every declared chart, reading a hidden block of live references."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    charts = [(s, c) for s in ir.sections for c in s.charts if s.table is not None]
    if not charts:
        return

    data_sheet = workbook.create_sheet(title=_CHART_DATA)
    data_sheet.sheet_state = "hidden"
    cursor = 1

    for section, chart in charts:
        table = section.table
        last_row, series_count = _chart_block(data_sheet, chart, table, layout, at_row=cursor)
        constructor = {"BarChart": BarChart, "LineChart": LineChart, "PieChart": PieChart}[
            _KIND[chart.kind.value][0]
        ]
        drawn = constructor()
        direction = _KIND[chart.kind.value][1]
        if direction is not None:
            drawn.type = direction
        drawn.title = chart.title
        if chart.category_axis:
            drawn.x_axis.title = chart.category_axis
        if chart.value_axis:
            drawn.y_axis.title = chart.value_axis

        drawn.add_data(
            Reference(data_sheet, min_col=2, max_col=1 + series_count,
                      min_row=cursor, max_row=last_row),
            titles_from_data=True,
        )
        drawn.set_categories(
            Reference(data_sheet, min_col=1, min_row=cursor + 1, max_row=last_row)
        )
        drawn.height, drawn.width = 8, 18

        # Anchored beside the table it plots, clear of the widest column.
        target = workbook[sheet_of[chart.table]]
        anchor = f"{_column_letter(len(table.columns) + 3)}{_HEADER_ROW + 1 + 18 * section.charts.index(chart)}"
        target.add_chart(drawn, anchor)

        cursor = last_row + 2

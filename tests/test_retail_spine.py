"""Retail's own organisation is load-bearing, or it is scenery.

The measurement this module holds. Retail was the vertical banking, insurance and
procurement were measured *against* — the one estate in the repository whose
stores and categories reached a compiled document — and it was also the one the
wave that made the other three load-bearing did not touch. So it ended that wave
as the laggard on exactly the dimension it had been the reference for, and
`validate.reachability` said so on one-period builds at seed 8128:

    archetype:omnichannel_retailer    cost centre 2/2
    archetype:australian_grocery      cost centre 2/2, site 44/1607
    pack:trading-retailer.json        cost centre 2/2, site  5/173
    pack:regional-insurer.json        cost centre 2/2

The last row is not a stray. `examples/packs/regional-insurer.json` declares
`"base": "retail"` — a general insurer described in the month-end close's own
vocabulary — so its cost centres were unreached for retail's reason and not for
insurance's, which is why `longtail-insurer.json` on the actual insurance engine
was clean beside it. `test_the_retail_engine_closes_every_company_it_builds`
below pins all four in one parametrised claim, because the interesting property
is that it is the *same sentence* for an archetype and for a pack.

`generators/retail_estate.py` is what closed them and argues the modelling. These
are the tests that stop it reopening, and they assert four things and nowhere
else:

* every declared cost centre and every declared site is **named by a fact** —
  the measurement itself, as a test;
* centres sum to the group and distribution centres to their division
  **exactly**, with `==` on integers and no tolerance, because the allocation is
  largest-remainder and a rounding note would be a defect;
* a **rendered** workbook carries the cost centre names and the warehouse rows —
  minting facts no artifact reports is this wave's defect one layer along, and an
  IR is not a file, so these load the `.xlsx` with `openpyxl` and read the cells;
* the two rates are **never summed**, which is the rule `columns.not_summable`
  and `Sheet.rate_kinds` state and this repository has already paid for twice.
"""

from __future__ import annotations

import io
from dataclasses import replace

import openpyxl
import pytest

from worldloom import World, packs, registries
from worldloom.generators import retail_estate
from worldloom.retail import RetailWorld
from worldloom.scenarios import MonthEndClose

SEED = 8128
PERIOD = "2026-03"


@pytest.fixture(scope="module")
def grocer() -> World:
    """The archetype with a distribution estate — 44 warehouses of 1,607 sites.

    `omnichannel_retailer` declares no warehouses at all, so it exercises the
    cost centres and nothing about the network. Both are asserted below; only
    this one can carry the site half.
    """
    from worldloom import archetypes

    return (
        RetailWorld(seed=SEED, archetype=archetypes.get("australian_grocery"))
        .build()
        .run(MonthEndClose(period=PERIOD))
    )


def _stated(world: World, kind: str) -> dict[str, float]:
    return {
        fact.subject: fact.value.amount
        for fact in world.facts.where(kind=kind)
        if fact.value and not fact.is_superseded
    }


# ---------------------------------------------------------------------------
# The measurement
# ---------------------------------------------------------------------------


#: Every way this repository builds a retail company, and what each one refused.
#: Two archetypes and two packs, because the engine fix has to generalise past
#: the archetype it was written against — that is the property the previous
#: wave's banking work established and the one a pack most easily breaks.
COMPANIES: dict[str, tuple[str, str]] = {
    "omnichannel_retailer": ("archetype", "omnichannel_retailer"),
    "australian_grocery": ("archetype", "australian_grocery"),
    "trading-retailer": ("pack", "examples/packs/trading-retailer.json"),
    # A retail-engine pack that describes an insurer. It is here precisely
    # because it looks like somebody else's problem and is not.
    "regional-insurer": ("pack", "examples/packs/regional-insurer.json"),
}


def _built(kind: str, source: str) -> World:
    from worldloom import archetypes

    if kind == "archetype":
        world = RetailWorld(seed=SEED, archetype=archetypes.get(source)).build()
    else:
        world = RetailWorld.from_pack(packs.load(source), seed=SEED).build()
    return world.run(MonthEndClose(period=PERIOD))


@pytest.mark.parametrize("company", sorted(COMPANIES))
def test_the_retail_engine_closes_every_company_it_builds(company: str) -> None:
    """The gate, over the whole engine rather than over one archetype.

    A fix that reached `omnichannel_retailer` and not `trading-retailer.json`
    would be a fix in the archetype table wearing a generator's clothes, which
    is the distinction the previous wave's measurement drew and the reason it
    was run across packs at all.

    Sites are asserted through the compiled world rather than through
    `validate.reachability`, deliberately: that function is the gate and lives
    in `validate.py`, and a spine test that could only fail when the gate was
    imported would be testing the gate. This asserts the corpus.

    Wrapped in `registries.scoped()` because two of these companies are packs
    and `packs.archetype_of` installs their authored document types into
    process-global tables that no build ever un-installs — deliberately, since
    a world is compiled long after it is built. Reading a pack here without a
    scope leaves `trade_pipeline_review` declared for every test that runs
    afterwards, which `tests/test_registries.py` then correctly refuses. The
    world is only read for its facts inside the block, so nothing it holds
    needs the installed tables once the block exits.
    """
    with registries.scoped():
        world = _built(*COMPANIES[company])
        named = {fact.subject for fact in world.facts}
        centres = {c.id for c in world.cost_centres}
        sites = {s.id for s in world.sites}

    missing_centres = sorted(centres - named)
    assert not missing_centres, (
        f"{company}: {len(missing_centres)} cost centre(s) are named by no fact."
        " A cost centre is where a charge lands; a company that declares two and"
        " books to neither has an accounting structure with no accounting in it."
    )

    missing_sites = sorted(sites - named)
    assert not missing_sites, (
        f"{company}: {len(missing_sites)} site(s) are named by no fact:"
        f" {missing_sites[:5]}"
    )


def test_the_warehouses_are_the_sites_the_store_p_and_l_leaves_out(grocer: World) -> None:
    """The two populations partition the estate, and neither is empty.

    This is what makes the fix a closure rather than a move. `finance.generate`
    drops zero-weight sites from the store P&L — correctly: a warehouse given
    turnover would reconcile to the unit total and still be nonsense — and
    `retail_estate` picks up exactly those and nothing else. Every site is
    therefore measured, and no site carries a measure its format cannot own.
    """
    revenue = _stated(grocer, "financial.revenue.actual")
    throughput = _stated(grocer, retail_estate.THROUGHPUT)

    warehouses = {s.id for s in retail_estate.distribution_estate(list(grocer.sites))}
    stores = {s.id for s in grocer.sites} - warehouses
    assert len(warehouses) == 44, len(warehouses)
    assert stores

    assert warehouses <= set(throughput)
    assert not (warehouses & set(revenue)), (
        "a warehouse booked turnover — the store P&L's own rule, broken"
    )
    assert stores <= set(revenue)
    assert not (stores & set(throughput)), (
        "a trading store was given a distribution throughput; this module cuts"
        " the network by the sites that *are* the network"
    )


# ---------------------------------------------------------------------------
# Reconciliation, exactly
# ---------------------------------------------------------------------------


def test_the_corporate_base_is_decomposed_two_ways(grocer: World) -> None:
    """The centres that incur the cost and the divisions that carry it are
    different sets of entities summing to one number, which is what makes
    either of them checkable."""
    cost = _stated(grocer, retail_estate.SHARED_COST)
    recharge = _stated(grocer, retail_estate.SHARED_RECHARGE)
    group = cost[grocer.company.id]

    centres = [cost[c.id] for c in grocer.cost_centres]
    assert len(centres) == len(list(grocer.cost_centres)) == 2
    # `==`, not `RECONCILIATION_TOLERANCE`: `finance.allocate` is
    # largest-remainder from a total drawn once, so the integer parts add to the
    # integer whole or the allocator is broken. A tolerance here would absorb
    # exactly the drift a later slide back to round-and-hope would introduce.
    assert sum(centres) == group

    carried = [recharge[u.id] for u in grocer.business_units if u.id in recharge]
    assert len(carried) == len(list(grocer.business_units))
    assert sum(carried) == group

    # And the two decompositions are over genuinely different entities, or the
    # cross-check is a relabelling.
    assert {c.id for c in grocer.cost_centres} & {u.id for u in grocer.business_units} == set()


@pytest.mark.parametrize(
    "kind", (retail_estate.THROUGHPUT, retail_estate.COST_TO_SERVE))
def test_distribution_centres_sum_to_their_division_exactly(
        grocer: World, kind: str) -> None:
    stated = _stated(grocer, kind)
    checked = 0
    for unit in grocer.business_units:
        estate = retail_estate.distribution_estate(
            [s for s in grocer.sites if s.business_unit_id == unit.id])
        parts = [stated[s.id] for s in estate if s.id in stated]
        if not parts or unit.id not in stated:
            continue
        checked += 1
        assert sum(parts) == stated[unit.id], f"{kind} for {unit.name}"
    assert checked, f"no division decomposed {kind} — the network is not reported"


@pytest.mark.parametrize(
    "kind", (retail_estate.THROUGHPUT, retail_estate.COST_TO_SERVE))
def test_the_network_total_is_the_divisions_that_have_one(
        grocer: World, kind: str) -> None:
    """`documents._sum_row`'s rule, stated over facts.

    The group figure is the sum of the divisions with a distribution estate, not
    a figure drawn for the company — so a corpus whose digital arm ships from a
    third party states a network total its own rows add up to.
    """
    stated = _stated(grocer, kind)
    divisions = [stated[u.id] for u in grocer.business_units if u.id in stated]
    assert divisions
    assert sum(divisions) == stated[grocer.company.id]


def test_a_company_with_no_warehouses_states_no_network_at_all() -> None:
    """The honest empty, and the reason an absent tab is not a missing one.

    `omnichannel_retailer` declares no zero-weight site, so there is no network
    to report and nothing is minted — `documents.finance_workbook`'s "an empty
    Store Performance tab is worse than no tab", one sheet along. Its cost
    centres still close, which is what makes this a separate claim rather than
    a weaker version of the one above.
    """
    world = _built("archetype", "omnichannel_retailer")
    assert not retail_estate.distribution_estate(list(world.sites))
    assert not _stated(world, retail_estate.THROUGHPUT)
    assert len(_stated(world, retail_estate.SHARED_COST)) == 3  # group + two centres


# ---------------------------------------------------------------------------
# A rate is not a total
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "rate, numerator, denominator, scale",
    (
        (retail_estate.SHARED_RECOVERY, retail_estate.SHARED_RECHARGE,
         "financial.revenue.actual", 100.0),
        (retail_estate.COST_PER_CARTON, retail_estate.COST_TO_SERVE,
         retail_estate.THROUGHPUT, 1.0),
    ),
)
def test_a_rate_is_derived_and_never_summed(
        grocer: World, rate: str, numerator: str, denominator: str, scale: float) -> None:
    stated = _stated(grocer, rate)
    top = _stated(grocer, numerator)
    bottom = _stated(grocer, denominator)

    for subject, value in stated.items():
        assert value == round(top[subject] * scale / bottom[subject], 2), subject

    divisions = [stated[u.id] for u in grocer.business_units if u.id in stated]
    assert len(divisions) >= 2, "one division cannot demonstrate the rule"
    # `columns.not_summable`, on real figures: the group rate is the group's own
    # two amounts divided, and the total of the divisional rates is a number
    # that means nothing. The strict inequalities are the stronger half — they
    # fail on a recharge basis that makes every division recover at the same
    # percentage, which is what a pure turnover allocation did before
    # `retail.overhead.service_intensity` existed.
    group = stated[grocer.company.id]
    assert group != sum(divisions)
    assert min(divisions) < group < max(divisions)


# ---------------------------------------------------------------------------
# A rendered artifact carries them
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def workbook(grocer: World):  # type: ignore[no-untyped-def]
    rendered = grocer.render("xlsx")
    item = next(
        r for r in rendered._rendered
        if r.path.endswith(".xlsx") and "month-end-model" in r.path
    )
    return openpyxl.load_workbook(io.BytesIO(item.payload))


def test_the_rendered_workbook_names_every_cost_centre_beside_a_number(
        grocer: World, workbook) -> None:  # type: ignore[no-untyped-def]
    """The assertion this whole lane turns on.

    Minting cost-centre facts that no artifact carries reproduces the defect it
    is supposed to close, one layer along, and the supporting-fact appendix
    would satisfy a weaker check while showing a reader nothing — see
    `validate._readable_surface`. So this loads the file and requires each
    centre's *name* to appear on a row that also carries a number.
    """
    sheet = workbook["Corporate Cost Base"]
    printed = {
        row[0]: row[1]
        for row in sheet.iter_rows(values_only=True)
        if row and isinstance(row[0], str)
    }
    cost = _stated(grocer, retail_estate.SHARED_COST)
    for centre in grocer.cost_centres:
        assert centre.name in printed, f"{centre.name} is on no sheet of the model"
        assert printed[centre.name] == cost[centre.id], centre.name
    # The total is a declared sum over the rows above it, not a pasted figure:
    # a reader who deletes a centre must see the total move.
    assert str(printed["Total corporate cost"]).startswith("=SUM(")


def test_the_rendered_workbook_reports_every_warehouse(
        grocer: World, workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Distribution Network"]
    header = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    cartons = header.index("Cartons dispatched")
    cost_column = header.index("Cost to serve")

    by_name = {site.name: site.id for site in grocer.sites}
    throughput = _stated(grocer, retail_estate.THROUGHPUT)
    cost = _stated(grocer, retail_estate.COST_TO_SERVE)

    checked = 0
    for row in sheet.iter_rows(min_row=4, values_only=True):
        site_id = by_name.get(row[0]) if row and isinstance(row[0], str) else None
        if site_id is None or site_id not in throughput:
            continue
        checked += 1
        assert row[cartons] == throughput[site_id], row[0]
        assert row[cost_column] == cost[site_id], row[0]
    assert checked == 44, f"{checked} of 44 warehouses on the sheet"


def test_the_recharge_sheet_lets_a_reader_recompute_the_rate(workbook) -> None:  # type: ignore[no-untyped-def]
    """Recovery is the two columns beside it divided, on the sheet itself.

    The revenue column is on this tab for exactly this reason: a recovery
    percentage a reader has to open another tab to check is a figure they will
    not check.
    """
    sheet = workbook["Shared Services Recharge"]
    header = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    recharge, revenue, recovery = (
        header.index("Recharge"), header.index("Revenue"), header.index("Recovery (%)"))

    checked = 0
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row or not isinstance(row[recharge], (int, float)):
            continue  # the Group row's money cells are formulas
        checked += 1
        assert row[recovery] == round(row[recharge] / row[revenue] * 100, 2), row[0]
    assert checked >= 3, "fewer divisions than the grocer declares"


def test_the_model_carries_every_fact_it_was_planned_to(grocer: World) -> None:
    """`carried_evidence`, for this artifact specifically: the plan's
    `required_fact_ids` against the compiled document's own."""
    compiled = grocer.compile()
    intent = next(
        i for i in compiled.artifact_intents if i.artifact_type == "finance_workbook"
    )
    ir = next(ir for ir in compiled.artifact_irs if ir.intent_id == intent.id)
    assert not set(intent.required_fact_ids) - set(ir.fact_ids())


# ---------------------------------------------------------------------------
# The check can fail
# ---------------------------------------------------------------------------


def test_a_broken_recharge_is_refused(grocer: World) -> None:
    """A check that passes on a corpus nobody broke is a check that is not
    looking. This moves one division's recharge by three and reads the verdict."""
    from worldloom.retail import _checks

    unit = next(iter(grocer.business_units))
    broken = [
        fact.model_copy(update={
            "value": fact.value.model_copy(update={"amount": fact.value.amount + 3})
        })
        if fact.kind == retail_estate.SHARED_RECHARGE and fact.subject == unit.id
        else fact
        for fact in grocer.facts
    ]
    violations, checks = _checks(replace(grocer, _facts=tuple(broken)))
    assert checks > 0
    assert "shared_services_recharge_does_not_reconcile" in {v.code for v in violations}


def test_a_broken_network_roll_up_is_refused(grocer: World) -> None:
    from worldloom.retail import _checks

    site = next(iter(retail_estate.distribution_estate(list(grocer.sites))))
    broken = [
        fact.model_copy(update={
            "value": fact.value.model_copy(update={"amount": fact.value.amount + 7})
        })
        if fact.kind == retail_estate.THROUGHPUT and fact.subject == site.id
        else fact
        for fact in grocer.facts
    ]
    violations, _ = _checks(replace(grocer, _facts=tuple(broken)))
    assert "distribution_network_does_not_reconcile" in {v.code for v in violations}


def test_a_rate_that_disagrees_with_its_amounts_is_refused(grocer: World) -> None:
    from worldloom.retail import _checks

    unit = next(iter(grocer.business_units))
    broken = [
        fact.model_copy(update={
            "value": fact.value.model_copy(update={"amount": fact.value.amount + 1.5})
        })
        if fact.kind == retail_estate.SHARED_RECOVERY and fact.subject == unit.id
        else fact
        for fact in grocer.facts
    ]
    violations, _ = _checks(replace(grocer, _facts=tuple(broken)))
    assert "rate_disagrees_with_its_amounts" in {v.code for v in violations}


# ---------------------------------------------------------------------------
# Three periods
# ---------------------------------------------------------------------------


def test_three_closes_validate_and_each_has_its_own_network() -> None:
    """`--periods 3`, and the figures move with the month.

    The draws are keyed on streams derived per scenario run, so a flat series
    would mean the period never reached the rng — the same claim
    `test_banking_spine` makes about its quarters.
    """
    from worldloom import archetypes

    built = RetailWorld(
        seed=SEED, archetype=archetypes.get("australian_grocery")).build()
    periods = ("2026-01", "2026-02", "2026-03")
    for period in periods:
        built = built.run(MonthEndClose(period=period))

    report = built.validate()
    assert report.ok, "\n".join(str(v) for v in report.violations)

    group = [
        fact.value.amount
        for fact in built.facts.where(kind=retail_estate.SHARED_COST)
        if fact.subject == built.company.id
    ]
    assert len(group) == len(periods)
    assert len(set(group)) == len(periods)

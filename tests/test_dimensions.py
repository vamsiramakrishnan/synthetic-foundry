"""Category and store dimensions: does a large world still reconcile?

Three divisions and a group row reconcile by accident. Thirty-four categories and
sixteen hundred stores, each allocated from a unit total and each summed back to
it on a different sheet, do not — so this module checks the arithmetic at every
level, and checks it by *evaluating the workbook's formulas* rather than by
re-reading the numbers the generator wrote.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from test_render import evaluate

from worldloom import MonthEndClose, RetailWorld
from worldloom.archetypes import AUSTRALIAN_GROCERY, available, get, inspired_by
from worldloom.generators.finance import allocate, previous_periods
from worldloom.world import World

PERIOD = "2026-03"
COMPARATIVES = 3


@pytest.fixture(scope="module")
def grocery() -> World:
    world = RetailWorld(seed=8128, archetype=AUSTRALIAN_GROCERY).build()
    world = world.run(
        MonthEndClose(
            period=PERIOD,
            include_operational_incident=True,
            comparative_months=COMPARATIVES,
        )
    )
    return world.render("xlsx")


@pytest.fixture(scope="module")
def book(grocery: World):  # type: ignore[no-untyped-def]
    item = next(r for r in grocery._rendered if r.path.endswith(".xlsx"))
    return openpyxl.load_workbook(io.BytesIO(item.payload))


# ---------------------------------------------------------------------------
# Allocation
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "total,weights",
    [
        (1_000_000, [0.3, 0.3, 0.4]),
        (7, [1, 1, 1]),  # fewer units than rows: the remainder has to land somewhere
        (1, [0.5, 0.5]),
        (-45_812, [0.118, 0.104, 0.778]),  # a variance is negative and still exact
        (999_999_999, [1.0] * 137),
    ],
)
def test_allocation_is_exact(total: int, weights: list[float]) -> None:
    """Largest-remainder, so the parts add to the whole with no residual line."""
    parts = allocate(total, weights)
    assert sum(parts) == total
    assert len(parts) == len(weights)


def test_allocation_is_proportional() -> None:
    """Exactness must not come at the cost of putting everything in one row."""
    parts = allocate(1_000_000, [0.5, 0.3, 0.2])
    assert parts == [500_000, 300_000, 200_000]


def test_allocation_refuses_a_degenerate_split() -> None:
    with pytest.raises(ValueError):
        allocate(100, [0.0, 0.0])


def test_previous_periods_walks_back_across_a_year() -> None:
    assert previous_periods("2026-02", 3) == ("2025-11", "2025-12", "2026-01")
    assert previous_periods("2026-03", 0) == ()


# ---------------------------------------------------------------------------
# The dimensions themselves
# ---------------------------------------------------------------------------


def test_the_hierarchy_is_populated(grocery: World) -> None:
    assert len(grocery.categories) == AUSTRALIAN_GROCERY.category_count
    assert len(grocery.sites) == AUSTRALIAN_GROCERY.site_count
    assert len(grocery.business_units) == len(AUSTRALIAN_GROCERY.units)


def test_every_category_has_a_unit_and_an_accountable_buyer(grocery: World) -> None:
    """A category is a noun someone owns. An unowned one cannot be argued about."""
    units = set(grocery.business_units.ids())
    people = set(grocery.people.ids())
    for category in grocery.categories:
        assert category.business_unit_id in units
        assert category.buyer_id in people


def test_distribution_centres_book_no_revenue(grocery: World) -> None:
    """A warehouse that reconciled would still be wrong."""
    booked = {f.subject for f in grocery.facts if f.kind == "financial.revenue.actual"}
    warehouses = [s for s in grocery.sites if "Distribution Centre" in s.format]
    assert warehouses, "the archetype should have distribution centres to exclude"
    for site in warehouses:
        assert site.revenue_weight == 0.0
        assert site.id not in booked


def test_a_large_world_validates(grocery: World) -> None:
    report = grocery.validate()
    assert report.ok, report.violations[:5]
    # The roll-up checks are the expensive ones; make sure they actually ran.
    assert report.checks_run > 20_000


def test_the_trend_covers_every_generated_period(grocery: World) -> None:
    periods = sorted({f.period for f in grocery.facts if f.period})
    assert len(periods) == COMPARATIVES + 1
    assert periods[-1] == PERIOD


def test_prior_periods_carry_actuals_but_not_budgets(grocery: World) -> None:
    """A trend needs actuals. Generating budgets nobody reads would treble the ledger."""
    past = previous_periods(PERIOD, 1)[0]
    kinds = {f.kind for f in grocery.facts if f.period == past}
    assert "financial.revenue.actual" in kinds
    assert "financial.revenue.budget" not in kinds


def test_prior_period_facts_predate_the_close(grocery: World) -> None:
    """A figure known before the close must be dated before it, or nothing may cite it."""
    close = max(f.valid_from for f in grocery.facts if f.period == PERIOD)
    for fact in grocery.facts:
        if fact.period and fact.period < PERIOD:
            assert fact.valid_from < close


# ---------------------------------------------------------------------------
# Reconciliation, at every level
# ---------------------------------------------------------------------------


def _unit_total(world: World, kind: str, subject: str) -> float:
    fact = next(
        f for f in world.facts
        if f.kind == kind and f.subject == subject and f.period == PERIOD
    )
    return fact.value.amount


def test_categories_and_stores_reach_the_same_unit_total(grocery: World) -> None:
    """Two independent decompositions of one number.

    This is the property the whole dimensional model rests on: a category P&L and
    a store P&L are different cuts of the same month, and a corpus where they
    disagree has two answers to one question.
    """
    for unit in grocery.business_units:
        stated = _unit_total(grocery, "financial.revenue.actual", unit.id)
        by_category = sum(
            f.value.amount
            for f in grocery.facts
            if f.kind == "financial.revenue.actual"
            and f.period == PERIOD
            and f.subject in {c.id for c in grocery.categories if c.business_unit_id == unit.id}
        )
        by_site = sum(
            f.value.amount
            for f in grocery.facts
            if f.kind == "financial.revenue.actual"
            and f.period == PERIOD
            and f.subject in {s.id for s in grocery.sites if s.business_unit_id == unit.id}
        )
        assert by_category == stated, f"{unit.name}: categories do not sum to the unit"
        assert by_site == stated, f"{unit.name}: stores do not sum to the unit"


def test_a_broken_category_allocation_is_caught(grocery: World) -> None:
    """The validator has to be able to fail, or it is decoration.

    Corrupt one category's revenue by a single unit and the category roll-up must
    report it — the unit and group figures are untouched, so nothing else moves.
    """
    facts = list(grocery._facts)
    index = next(i for i, f in enumerate(facts) if f.subject.startswith("CAT-") and f.value)
    facts[index] = facts[index].model_copy(
        update={"value": facts[index].value.model_copy(
            update={"amount": facts[index].value.amount + 1_000})}
    )
    broken = World(**{**grocery.__dict__, "_facts": tuple(facts)})

    report = broken.validate()
    assert not report.ok
    assert any(v.code == "does_not_reconcile" for v in report.violations)


def test_a_broken_store_allocation_is_caught(grocery: World) -> None:
    facts = list(grocery._facts)
    index = next(i for i, f in enumerate(facts) if f.subject.startswith("SITE-") and f.value)
    facts[index] = facts[index].model_copy(
        update={"value": facts[index].value.model_copy(
            update={"amount": facts[index].value.amount + 5_000})}
    )
    broken = World(**{**grocery.__dict__, "_facts": tuple(facts)})

    assert any(v.code == "does_not_reconcile" for v in broken.validate().violations)


# ---------------------------------------------------------------------------
# The workbook
# ---------------------------------------------------------------------------


def test_the_workbook_has_the_dimensional_sheets(book) -> None:  # type: ignore[no-untyped-def]
    titles = [sheet.title for sheet in book.worksheets]
    assert "Category P&L" in titles
    assert "Store Performance" in titles
    assert "Revenue Trend" in titles


def test_the_workbook_is_actually_large(book) -> None:
    """The point of the exercise. A proof of mechanism is not a workbook."""
    rows = sum(sheet.max_row for sheet in book.worksheets)
    assert rows > 5_000, f"only {rows} rows across {len(book.worksheets)} sheets"
    assert book["Store Performance"].max_row > 1_500


def test_a_subtotal_sums_its_own_children_and_nothing_else(book) -> None:  # type: ignore[no-untyped-def]
    """The bug a contiguous-range shortcut would cause.

    On the category sheet the group row's operands are the unit subtotals, which
    are scattered down the sheet. Emitted as ``=SUM(first:last)`` that range also
    swallows every category row in between, and the group total doubles. So the
    group row must be a list, and the unit subtotals — which really are
    consecutive — must be a range.
    """
    sheet = book["Category P&L"]
    group_row = next(
        row for row in range(4, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value == "Group"
    )
    group_formula = sheet.cell(row=group_row, column=3).value
    assert group_formula.startswith("=SUM(")
    assert "," in group_formula, f"scattered operands collapsed into a range: {group_formula}"
    assert ":" not in group_formula

    subtotal_row = next(
        row for row in range(4, group_row)
        if str(sheet.cell(row=row, column=1).value or "").endswith(" total")
    )
    assert ":" in sheet.cell(row=subtotal_row, column=3).value


def test_every_workbook_formula_evaluates_to_its_fact(grocery: World, book) -> None:  # type: ignore[no-untyped-def]
    """Recompute the sheet and compare against the ledger, cell by cell.

    The store sheet alone is several thousand formulas. openpyxl stores them
    without evaluating, so nothing else in this suite would notice a renderer that
    emitted the wrong range.
    """
    ir = next(r for r in grocery.artifact_irs if r.tables() and r.tables()[0].key == "summary")
    checked = 0

    for section in ir.sections:
        table = section.table
        if table is None or table.key in ("lineage", "drivers", "incident_impact"):
            continue
        sheet_name = table.title[:31]
        columns = {column.key: index for index, column in enumerate(table.columns)}
        for row_index, row in enumerate(table.rows):
            for column_key, cell in row.cells.items():
                if cell.formula is None or not isinstance(cell.value, (int, float)):
                    continue
                column = table.columns[columns[column_key]]
                address = f"{openpyxl.utils.get_column_letter(columns[column_key] + 2)}{4 + row_index}"
                computed = evaluate(book, sheet_name, address)
                expected = cell.value / 100 if column.number_format == "0.00%" else cell.value
                assert abs(computed - expected) < 0.01, (
                    f"{sheet_name}!{address} ({row.label}/{column_key}): "
                    f"formula gives {computed}, ledger says {expected}"
                )
                checked += 1

    assert checked > 500, f"only checked {checked} formulas"


def test_the_reconciliation_sheet_checks_every_level(grocery: World, book) -> None:  # type: ignore[no-untyped-def]
    """Units to group, categories to unit, stores to unit — all netting to zero."""
    sheet = book["Reconciliation"]
    rows = [r for r in range(4, sheet.max_row + 1) if sheet.cell(row=r, column=2).value is not None]
    labels = [sheet.cell(row=r, column=1).value for r in rows]

    assert any("categories sum to the unit" in str(label) for label in labels)
    assert any("stores sum to the unit" in str(label) for label in labels)
    assert len(rows) >= 2 + 2 * len(grocery.business_units)

    for row in rows:
        difference = evaluate(book, "Reconciliation", f"D{row}")
        assert abs(difference) < 1.0, f"{labels[rows.index(row)]} does not net to zero: {difference}"


def test_a_corrupted_stated_total_moves_the_reconciliation(grocery: World) -> None:
    """The check compares against the ledger, not against the sheet's own sum.

    If it compared the summed units against the P&L's group cell — itself
    ``=SUM(units)`` — it could never disagree, and would prove nothing.
    """
    facts = list(grocery._facts)
    index = next(
        i for i, f in enumerate(facts)
        if f.kind == "financial.revenue.actual"
        and f.subject == grocery.company.id
        and f.period == PERIOD
    )
    facts[index] = facts[index].model_copy(
        update={"value": facts[index].value.model_copy(
            update={"amount": facts[index].value.amount + 9_000})}
    )
    broken = World(**{**grocery.__dict__, "_facts": tuple(facts), "_artifact_irs": ()})
    rendered = broken.render("xlsx")

    item = next(r for r in rendered._rendered if r.path.endswith(".xlsx"))
    sheet = openpyxl.load_workbook(io.BytesIO(item.payload))
    difference = evaluate(sheet, "Reconciliation", "D4")
    assert abs(difference + 9_000) < 1.0, f"the check did not notice the corrupted total: {difference}"


# ---------------------------------------------------------------------------
# Archetypes
# ---------------------------------------------------------------------------


def test_inspired_by_matches_the_longest_phrase() -> None:
    assert inspired_by("woolies").key == "australian_grocery"
    assert inspired_by("a large Australian retailer like Woolworths").key == "australian_grocery"
    assert inspired_by("an omnichannel retailer").key == "omnichannel_retailer"


def test_inspired_by_falls_back_rather_than_raising() -> None:
    """An unrecognised description should get a working world, not an error."""
    assert inspired_by("a regional widget conglomerate").key in available()


def test_an_inspired_world_borrows_shape_and_no_data() -> None:
    """The whole point of `inspired_by`: the same kind of business, wholly invented."""
    builder = RetailWorld.inspired_by("a large Australian grocer like Woolworths", seed=8128)
    assert builder.archetype.key == "australian_grocery"

    world = builder.build()
    haystack = " ".join(
        [world.company.name, world.company.headquarters]
        + [unit.name for unit in world.business_units]
        + [person.name for person in world.people]
        + [site.name for site in world.sites]
    ).casefold()
    for banned in ("woolworth", "woolies", "coles", "countdown", "big w"):
        assert banned not in haystack


def test_unknown_archetype_names_the_ones_that_exist() -> None:
    with pytest.raises(KeyError) as caught:
        get("not_a_shape")
    assert "omnichannel_retailer" in str(caught.value)

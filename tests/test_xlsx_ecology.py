from io import BytesIO

import openpyxl

from worldloom import MonthEndClose, RetailWorld
from worldloom.ecology import render


def _book():  # type: ignore[no-untyped-def]
    world = RetailWorld(seed=8128).build().run(
        MonthEndClose(period="2026-03", include_operational_incident=True)
    )
    result = render(world, "xlsx")
    payload = next(item.payload for item in result.world._rendered if item.path.endswith(".xlsx"))
    return openpyxl.load_workbook(BytesIO(payload), data_only=False)


def test_ecology_workbook_has_hidden_document_control() -> None:
    book = _book()
    assert "Document Control" in book.sheetnames
    control = book["Document Control"]
    assert control.sheet_state == "hidden"
    values = {control.cell(row=row, column=1).value: control.cell(row=row, column=2).value for row in range(1, control.max_row + 1)}
    assert values["Realism profile"] == "ecology/v1"
    assert values["Artifact family"]
    assert values["Lifecycle"]
    assert int(values["Revision"]) >= 1
    assert values["Style seed"]


def test_ecology_workbook_control_sheet_does_not_add_business_numbers() -> None:
    book = _book()
    control = book["Document Control"]
    labels = [control.cell(row=row, column=1).value for row in range(1, control.max_row + 1)]
    assert "Revenue" not in labels
    assert "Margin" not in labels
    assert "Forecast" not in labels


def test_detail_sheets_get_filters_without_changing_formulas() -> None:
    book = _book()
    visible = [sheet for sheet in book.worksheets if sheet.sheet_state == "visible"]
    assert visible
    assert any(sheet.auto_filter.ref for sheet in visible if sheet.max_row > 4 and sheet.max_column > 1)

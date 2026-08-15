"""The bank's organisation is load-bearing, or it is scenery.

The measurement this module holds. A one-period banking build produced **58
facts across 11 documents**, and its three business units, its 133 branches and
both its cost centres were named by no fact and carried by no document. Retail
produced 604 facts from one period *because* its estate was load-bearing. The
generator that closed the gap is ``generators/banking_network``; these are the
tests that stop it reopening.

Four things are asserted here and nowhere else:

* every declared unit, site and cost centre is **named by a fact** — the
  measurement itself, as a test;
* branches sum to their division and divisions to the group **exactly**, with
  ``==`` on integers rather than a tolerance, because the allocation is
  largest-remainder and a rounding note would be a defect;
* a **rendered** workbook carries the branch figures — minting facts no
  artifact reports is the same defect one layer along, and an IR is not a file;
* the reconciliation check **fails when the arithmetic is broken**, because a
  check that cannot fail is decoration.
"""

from __future__ import annotations

import io
from dataclasses import replace

import openpyxl
import pytest

from worldloom import World
from worldloom.banking import BankingWorld
from worldloom.banking_scenarios import QuarterlyCapitalReturn

PERIOD = "2026-03"

#: The measures that decompose group → division → branch. Every one adds up;
#: ``banking.loan_to_deposit_pct`` is deliberately not here.
ADDITIVE = (
    "banking.deposits.balance",
    "banking.lending.balance",
    "banking.lending.settled",
    "banking.net_operating_income",
    "banking.network.fte",
)


@pytest.fixture(scope="module")
def world() -> World:
    return BankingWorld(seed=8128).build().run(QuarterlyCapitalReturn(period=PERIOD))


def _stated(world: World, kind: str) -> dict[str, float]:
    return {
        fact.subject: fact.value.amount
        for fact in world.facts.where(kind=kind)
        if fact.value and not fact.is_superseded
    }


# ---------------------------------------------------------------------------
# The measurement
# ---------------------------------------------------------------------------


def test_every_declared_entity_of_the_spine_is_named_by_a_fact(world: World) -> None:
    """The gap this vertical's whole wave exists to close, stated as a test.

    Business units, sites and cost centres — the three classes that reached
    nothing. Categories and services already did, through the capital return
    and the incident chain, so they are not restated here.
    """
    named = {fact.subject for fact in world.facts}
    for label, ids in (
        ("business units", [u.id for u in world.business_units]),
        ("sites", [s.id for s in world.sites]),
        ("cost centres", [c.id for c in world.cost_centres]),
    ):
        missing = sorted(set(ids) - named)
        assert not missing, f"{len(missing)} {label} are named by no fact: {missing[:5]}"


def test_the_estate_is_the_reason_the_corpus_is_large(world: World) -> None:
    """58 facts before, and the difference is the branch network.

    Pinned as a floor rather than an equality: a later change that adds a
    measure should not have to edit this, and one that silently drops the
    estate must not pass.
    """
    network = [f for f in world.facts if f.kind.startswith("banking.")]
    assert len(network) > 600, "the branch network mints the bulk of a bank's facts"
    assert len(list(world.facts)) > 700


# ---------------------------------------------------------------------------
# Reconciliation, exactly
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("kind", ADDITIVE)
def test_branches_sum_to_their_division_exactly(world: World, kind: str) -> None:
    stated = _stated(world, kind)
    checked = 0
    for unit in world.business_units:
        estate = [s.id for s in world.sites if s.business_unit_id == unit.id]
        parts = [stated[site] for site in estate if site in stated]
        if not parts or unit.id not in stated:
            continue
        checked += 1
        # `==`, not a tolerance: `finance.allocate` is largest-remainder, so the
        # integer parts add to the integer whole or the allocator is broken.
        assert sum(parts) == stated[unit.id], f"{kind} for {unit.name}"
    assert checked, f"no division decomposed {kind} — the estate is not being reported"


@pytest.mark.parametrize("kind", ADDITIVE)
def test_divisions_sum_to_the_group_exactly(world: World, kind: str) -> None:
    stated = _stated(world, kind)
    parts = [stated[unit.id] for unit in world.business_units if unit.id in stated]
    assert parts, f"no division states {kind}"
    assert sum(parts) == stated[world.company.id], kind


def test_a_division_with_no_branches_holds_no_customer_balances(world: World) -> None:
    """Treasury and Markets earns income and gathers no deposits.

    The honest statement, and the one that makes the roll-up above meaningful:
    if every division held balances the sum would reconcile for a trivial
    reason. Retail's "the digital unit had no estate" rule, in a bank's
    vocabulary.
    """
    deposits = _stated(world, "banking.deposits.balance")
    income = _stated(world, "banking.net_operating_income")
    branchless = [
        unit for unit in world.business_units
        if not any(s.business_unit_id == unit.id and s.revenue_weight > 0
                   for s in world.sites)
    ]
    assert branchless, "the shipped ADI has a treasury desk with no branches"
    for unit in branchless:
        assert unit.id not in deposits
        assert unit.id in income


# ---------------------------------------------------------------------------
# A rate is not a total
# ---------------------------------------------------------------------------


def test_the_loan_to_deposit_ratio_is_derived_and_never_summed(world: World) -> None:
    stated = _stated(world, "banking.loan_to_deposit_pct")
    lending = _stated(world, "banking.lending.balance")
    deposits = _stated(world, "banking.deposits.balance")

    for subject, ratio in stated.items():
        assert ratio == round(lending[subject] / deposits[subject] * 100, 2), subject

    divisions = [stated[u.id] for u in world.business_units if u.id in stated]
    assert len(divisions) >= 2, "one division cannot demonstrate the rule"
    # The whole point of `columns.not_summable`: the group figure is a ratio of
    # the group's own amounts, and the total of the divisional ratios is a
    # number that means nothing.
    assert stated[world.company.id] != sum(divisions)
    assert min(divisions) < stated[world.company.id] < max(divisions)


def test_the_shared_services_base_is_decomposed_two_ways(world: World) -> None:
    """The cost centres that incur it and the divisions that carry it are
    different sets of entities summing to one number — which is what makes
    either of them checkable."""
    cost = _stated(world, "banking.shared_services_cost")
    recharge = _stated(world, "banking.shared_services_recharge")
    group = cost[world.company.id]

    centres = [cost[c.id] for c in world.cost_centres if c.id in cost]
    assert len(centres) == len(list(world.cost_centres))
    assert sum(centres) == group

    carried = [recharge[u.id] for u in world.business_units if u.id in recharge]
    assert len(carried) == len(list(world.business_units))
    assert sum(carried) == group


# ---------------------------------------------------------------------------
# A rendered artifact carries them
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def workbook(world: World):  # type: ignore[no-untyped-def]
    rendered = world.render("xlsx")
    item = next(
        r for r in rendered._rendered
        if r.path.endswith(".xlsx") and "divisional" in r.path
    )
    return openpyxl.load_workbook(io.BytesIO(item.payload))


def test_the_rendered_workbook_reports_every_branch(world: World, workbook) -> None:  # type: ignore[no-untyped-def]
    """Minting site facts no artifact carries is the defect this wave closes,
    one layer along. An IR is not a file, so this reads the file."""
    printed: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str):
                printed.add(row[0])

    for site in world.sites:
        assert site.name in printed, f"{site.name} is on no sheet of the pack"
    for unit in world.business_units:
        assert unit.name in printed
    for centre in world.cost_centres:
        assert centre.name in printed


def test_the_branch_figures_on_the_sheet_are_the_ledgers(world: World, workbook) -> None:
    deposits = _stated(world, "banking.deposits.balance")
    by_name = {site.name: site.id for site in world.sites}
    sheet = next(s for s in workbook.worksheets if "branch network" in s.title)
    header = [cell for cell in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))]
    column = header.index("Deposits")

    checked = 0
    for row in sheet.iter_rows(min_row=4, values_only=True):
        site_id = by_name.get(row[0]) if row and isinstance(row[0], str) else None
        if site_id is None or site_id not in deposits:
            continue
        checked += 1
        assert row[column] == deposits[site_id], row[0]
    assert checked >= 100, f"only {checked} branches checked — the sheet is not the estate"


def test_the_division_total_is_a_formula_over_its_branches(workbook) -> None:
    """A workbook, not a screenshot of one — `documents.finance_workbook`'s rule."""
    sheet = next(s for s in workbook.worksheets if "branch network" in s.title)
    total = next(row for row in sheet.iter_rows(values_only=True)
             if row and isinstance(row[0], str) and row[0].endswith(" total"))
    formulas = [cell for cell in total if isinstance(cell, str) and cell.startswith("=SUM(")]
    assert len(formulas) == 5, "every measure on the subtotal row must be a declared sum"


def test_the_pack_carries_every_fact_it_was_planned_to(world: World) -> None:
    """`carried_evidence`, for this artifact specifically: the plan's
    `required_fact_ids` against the compiled document's own."""
    compiled = world.compile()
    intent = next(
        i for i in compiled.artifact_intents
        if i.artifact_type == "divisional_performance_pack"
    )
    ir = next(ir for ir in compiled.artifact_irs if ir.intent_id == intent.id)
    assert not set(intent.required_fact_ids) - set(ir.fact_ids())


# ---------------------------------------------------------------------------
# The check can fail
# ---------------------------------------------------------------------------


def test_a_broken_roll_up_is_refused(world: World) -> None:
    """A check that passes on a corpus nobody broke is a check that is not
    looking. This breaks one branch's deposits and reads the verdict."""
    from worldloom.banking import _checks

    site = next(s for s in world.sites if s.revenue_weight > 0)
    broken = []
    for fact in world.facts:
        if fact.kind == "banking.deposits.balance" and fact.subject == site.id:
            fact = fact.model_copy(update={
                "value": fact.value.model_copy(update={"amount": fact.value.amount + 7})
            })
        broken.append(fact)

    violations, checks = _checks(replace(world, _facts=tuple(broken)))
    assert checks > 0
    codes = {v.code for v in violations}
    assert "network_does_not_reconcile" in codes, sorted(codes)


def test_a_broken_recharge_is_refused(world: World) -> None:
    from worldloom.banking import _checks

    unit = next(iter(world.business_units))
    broken = [
        fact.model_copy(update={
            "value": fact.value.model_copy(update={"amount": fact.value.amount + 3})
        })
        if fact.kind == "banking.shared_services_recharge" and fact.subject == unit.id
        else fact
        for fact in world.facts
    ]
    violations, _ = _checks(replace(world, _facts=tuple(broken)))
    assert "shared_services_recharge_does_not_reconcile" in {v.code for v in violations}


# ---------------------------------------------------------------------------
# Three quarters
# ---------------------------------------------------------------------------


def test_three_quarters_validate_and_each_has_its_own_network() -> None:
    """`--periods 3` on this vertical, which steps three months at a time."""
    built = BankingWorld(seed=8128).build()
    periods = ("2026-03", "2026-06", "2026-09")
    for period in periods:
        built = built.run(QuarterlyCapitalReturn(period=period))

    report = built.validate()
    assert report.ok, "\n".join(str(v) for v in report.violations)

    deposits = {
        (fact.period, fact.subject)
        for fact in built.facts.where(kind="banking.deposits.balance")
    }
    for period in periods:
        assert sum(1 for p, _ in deposits if p == period) > 100, period

    # Each quarter's own figure, not last quarter's copied forward: the draws
    # are keyed on streams derived per scenario run, so a flat series would
    # mean the period never reached the rng.
    group = [
        fact.value.amount
        for fact in built.facts.where(kind="banking.net_operating_income")
        if fact.subject == built.company.id
    ]
    assert len(set(group)) == len(periods)

"""Charts, declared in the IR rather than drawn by each renderer.

Same principle as formulas, for the same reason. A workbook, a document and a deck
all show "revenue by division"; if each renderer picked its own rows and series
they would be three charts of three different things wearing one title. The IR
declares which table, which rows, which series, and a renderer decides only how to
draw it.

The checks that matter here are the ones about *which rows* — a chart that
included subtotals alongside the rows they total would plot the same money twice,
and it would look perfectly fine.
"""

from __future__ import annotations

import pytest

from worldloom import MonthEndClose, RetailWorld, World
from worldloom.models import ChartKind, FormulaKind
from worldloom.render import markdown

PERIOD = "2026-03"


@pytest.fixture(scope="module")
def workbook() -> World:
    world = RetailWorld(seed=8128).build().run(
        MonthEndClose(period=PERIOD, include_operational_incident=True, comparative_months=3)
    )
    return world.compile()


def _ir(world: World):  # type: ignore[no-untyped-def]
    return next(ir for ir in world.artifact_irs if ir.tables() and ir.tables()[0].key == "summary")


def test_the_workbook_declares_charts(workbook: World) -> None:
    charts = {chart.key for chart in _ir(workbook).charts()}
    assert {"pnl_revenue", "pnl_margin", "category_variance", "trend_units"} <= charts


def test_every_chart_resolves_against_its_own_table(workbook: World) -> None:
    """The check the validator runs, asserted directly so its failure is legible."""
    ir = _ir(workbook)
    tables = {table.key: table for table in ir.tables()}
    for chart in ir.charts():
        table = tables[chart.table]
        columns = {column.key for column in table.columns}
        rows = {row.key for row in table.rows}
        assert set(chart.series) <= columns, chart.key
        assert set(chart.rows) <= rows, chart.key
        assert chart.series, f"{chart.key} plots nothing"


def test_a_chart_never_plots_a_total_beside_its_own_parts(workbook: World) -> None:
    """The failure that looks correct.

    A chart of four divisions plus the group puts one bar four times the height of
    the rest beside them; a chart of categories plus their unit subtotals plots the
    same money twice. Neither raises, and both look fine.

    Note that plotting a subtotal is not itself the error — the trend chart plots
    one line per division, and a division is a subtotal of its categories. The
    error is plotting a total *together with the rows that total into it*, which
    the IR can answer exactly: a summing cell names its children as operands.
    """
    ir = _ir(workbook)
    tables = {table.key: table for table in ir.tables()}
    for chart in ir.charts():
        table = tables[chart.table]
        plotted = set(chart.rows)
        for row in table.rows:
            if row.key not in plotted:
                continue
            children = {
                operand
                for cell in row.cells.values()
                if cell.formula is FormulaKind.SUM
                for operand in cell.operands
            }
            overlap = children & plotted
            assert not overlap, (
                f"{chart.key} plots {row.key} and its own parts {sorted(overlap)} — "
                "the same money twice"
            )


def test_a_line_chart_is_only_used_where_the_axis_is_ordered(workbook: World) -> None:
    """A line between unordered categories asserts a trend that does not exist."""
    for chart in _ir(workbook).charts():
        if chart.kind is ChartKind.LINE:
            assert chart.table == "trend", f"{chart.key} draws a line over {chart.table}"


def test_a_broken_chart_reference_is_caught(workbook: World) -> None:
    """The validator has to be able to fail, or it is decoration.

    A chart naming a column that does not exist draws an empty series rather than
    raising: the file opens, the chart is there, and it is blank.
    """
    ir = _ir(workbook)
    sections = list(ir.sections)
    index, section = next(
        (i, s) for i, s in enumerate(sections) if s.charts
    )
    broken = section.charts[0].model_copy(update={"series": ["no_such_column"]})
    sections[index] = section.model_copy(update={"charts": [broken]})
    damaged = World(
        **{**workbook.__dict__, "_artifact_irs": (ir.model_copy(update={"sections": sections}),)}
    )

    report = damaged.validate()
    assert any(v.code == "chart_series_missing" for v in report.violations), report.violations[:3]


def test_a_chart_over_a_missing_table_is_caught(workbook: World) -> None:
    ir = _ir(workbook)
    sections = list(ir.sections)
    index, section = next((i, s) for i, s in enumerate(sections) if s.charts)
    broken = section.charts[0].model_copy(update={"table": "no_such_table"})
    sections[index] = section.model_copy(update={"charts": [broken]})
    damaged = World(
        **{**workbook.__dict__, "_artifact_irs": (ir.model_copy(update={"sections": sections}),)}
    )

    assert any(v.code == "chart_table_missing" for v in damaged.validate().violations)


def test_markdown_names_the_chart_rather_than_approximating_it(workbook: World) -> None:
    """Markdown cannot draw one, and an ASCII approximation would be a second
    rendering of the same data that could disagree with the table above it."""
    body = markdown.render(_ir(workbook)).decode()
    for chart in _ir(workbook).charts():
        assert f"**Figure — {chart.title}**" in body


def test_a_section_without_charts_still_reads_correctly(workbook: World) -> None:
    """Regression: a `for` between the `elif` and the `else` binds the `else` to
    the loop, and Python runs a loop-else whenever the loop did not break — so
    every chartless section grew an "awaiting narrative" notice under its table."""
    body = markdown.render(_ir(workbook)).decode()
    assert "Awaiting narrative" not in body


# ---------------------------------------------------------------------------
# Native charts in the workbook
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def book(workbook: World):  # type: ignore[no-untyped-def]
    import io

    import openpyxl

    rendered = workbook.render("xlsx")
    item = next(r for r in rendered._rendered if r.path.endswith(".xlsx"))
    return openpyxl.load_workbook(io.BytesIO(item.payload))


def _value_cells(book):  # type: ignore[no-untyped-def]
    """Every data cell of the chart block, skipping each block's header row.

    A header row holds the series labels, which are text by design — asserting
    that they too are formulas would be asserting the wrong thing.
    """
    sheet = book["Chart Data"]
    headers = {
        r for r in range(1, sheet.max_row + 1)
        if isinstance(sheet.cell(row=r, column=1).value, str)
        and sheet.cell(row=r, column=2).value is not None
        and not str(sheet.cell(row=r, column=2).value).startswith("='")
    }
    return [
        cell
        for row in sheet.iter_rows(min_col=2)
        for cell in row
        if cell.value is not None and cell.row not in headers
    ]


def test_the_workbook_carries_native_charts(workbook: World, book) -> None:  # type: ignore[no-untyped-def]
    """Not an image of a chart — a chart object Excel will redraw."""
    drawn = {
        sheet.title: [type(c).__name__ for c in getattr(sheet, "_charts", [])]
        for sheet in book.worksheets
        if getattr(sheet, "_charts", [])
    }
    assert drawn["Business Unit P&L"] == ["BarChart", "BarChart"]
    assert drawn["Category P&L"] == ["BarChart"]
    assert drawn["Revenue Trend"] == ["LineChart"]


def test_chart_data_is_live_references_not_pasted_values(book) -> None:  # type: ignore[no-untyped-def]
    """The property that makes a chart checkable.

    A chart fed pasted numbers is a screenshot: change a figure and the chart
    keeps the old story. Every value cell in the block is a cross-sheet formula,
    so the chart moves when the sheet does.
    """
    sheet = book["Chart Data"]
    assert sheet.sheet_state == "hidden"

    values = 0
    for cell in _value_cells(book):
        assert isinstance(cell.value, str) and cell.value.startswith("='"), (
            f"{cell.coordinate} holds a pasted value: {cell.value!r}"
        )
        values += 1
    assert values > 15, f"only {values} charted cells"


def test_every_charted_cell_resolves_to_the_figure_it_plots(workbook: World, book) -> None:  # type: ignore[no-untyped-def]
    """Recompute the reference and compare it against the IR, cell by cell.

    A block of formulas pointing at the wrong column would look exactly like a
    block pointing at the right one.
    """
    from test_render import evaluate

    ir = _ir(workbook)
    tables = {table.key: table for table in ir.tables()}

    charted = set()
    for cell in _value_cells(book):
        computed = evaluate(book, "Chart Data", cell.coordinate)
        assert isinstance(computed, float)
        charted.add(round(computed, 4))

    # A percentage is held on the sheet as a fraction so Excel's percent format
    # does not render 24.73 as 2473%, so a charted 0.2473 is the IR's 24.73.
    available = set()
    for table in tables.values():
        for table_row in table.rows:
            for cell in table_row.cells.values():
                if isinstance(cell.value, (int, float)):
                    available.add(round(float(cell.value), 4))
                    available.add(round(float(cell.value) / 100, 4))

    missing = sorted(charted - available)
    assert not missing, f"charted values present in no table: {missing[:5]}"
    assert len(charted) > 10


def test_a_by_row_chart_plots_one_series_per_row(workbook: World, book) -> None:  # type: ignore[no-untyped-def]
    """The trend wants one line per division across months.

    Drawn the other way round it is one line per month across divisions — twelve
    lines of a single point each — and it renders without complaint.
    """
    ir = _ir(workbook)
    trend = next(c for c in ir.charts() if c.key == "trend_units")
    assert trend.by_row

    sheet = book["Chart Data"]
    header = next(
        r for r in range(1, sheet.max_row + 1)
        if sheet.cell(row=r, column=1).value == trend.title
    )
    series = [
        sheet.cell(row=header, column=c).value
        for c in range(2, sheet.max_column + 1)
        if sheet.cell(row=header, column=c).value
    ]
    categories = [
        sheet.cell(row=r, column=1).value
        for r in range(header + 1, header + 1 + len(trend.series))
    ]
    assert len(series) == len(trend.rows), "series should be divisions"
    assert categories == list(trend.series), "the axis should be the periods, in order"


def test_charts_do_not_break_byte_identical_rendering(workbook: World) -> None:
    from worldloom.render import xlsx

    ir = _ir(workbook)
    assert xlsx.render(ir) == xlsx.render(ir)

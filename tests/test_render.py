"""Step 5: renderers.

The interesting test is not that a file appears. It is that the workbook's
*formulas* evaluate to the facts they came from — openpyxl stores formulas without
computing them, so a renderer could emit plausible nonsense and every naive test
would pass. So this module evaluates them.
"""

from __future__ import annotations

import io
import json
import re

import openpyxl
import pytest

from worldloom import MonthEndClose, RetailWorld, World
from worldloom.render import RenderError, available, markdown, xlsx

PERIOD = "2026-03"


@pytest.fixture(scope="module")
def rendered() -> World:
    world = RetailWorld(seed=8128).build().run(
        MonthEndClose(period=PERIOD, include_operational_incident=True)
    )
    return world.render("xlsx", "markdown", "jira", "confluence", "servicenow")


@pytest.fixture(scope="module")
def workbook(rendered: World):  # type: ignore[no-untyped-def]
    item = next(r for r in rendered._rendered if r.path.endswith(".xlsx"))
    return openpyxl.load_workbook(io.BytesIO(item.payload))


# ---------------------------------------------------------------------------
# A minimal evaluator for the formula shapes the renderer emits
# ---------------------------------------------------------------------------

# Unnamed, so the pattern can be composed with itself without colliding groups.
_CELL = r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+"


def _split_cell(token: str) -> tuple[str | None, str]:
    """A cell token into ``(sheet or None, address)``."""
    token = token.strip().replace("$", "")
    if "!" in token:
        sheet, _, address = token.partition("!")
        return sheet.strip("'"), address
    return None, token


def evaluate(book, sheet_name: str, address: str, depth: int = 0) -> float:  # type: ignore[no-untyped-def]
    """Resolve a cell, computing the formula shapes the renderer is allowed to emit."""
    if depth > 12:
        raise AssertionError(f"formula recursion too deep at {sheet_name}!{address}")

    value = book[sheet_name][address.replace("$", "")].value
    if not isinstance(value, str) or not value.startswith("="):
        return float(value) if isinstance(value, (int, float)) else 0.0
    return _compute(book, sheet_name, value[1:], depth + 1)


def _resolve(book, sheet_name: str, token: str, depth: int) -> float:  # type: ignore[no-untyped-def]
    if not re.fullmatch(_CELL, token.strip()):
        return float(token)
    sheet, address = _split_cell(token)
    return evaluate(book, sheet or sheet_name, address, depth)


def _compute(book, sheet_name: str, expression: str, depth: int) -> float:  # type: ignore[no-untyped-def]
    expression = expression.strip()

    guard = re.fullmatch(rf"IF\(({_CELL})=0,0,({_CELL})/({_CELL})\)", expression)
    if guard:
        denominator = _resolve(book, sheet_name, guard.group(1), depth)
        if denominator == 0:
            return 0.0
        return _resolve(book, sheet_name, guard.group(2), depth) / denominator

    # SUM(range) - cell, which is the reconciliation shape.
    sum_then_minus = re.fullmatch(rf"SUM\((.+?)\)\s*-\s*({_CELL})", expression)
    if sum_then_minus:
        total = _sum(book, sheet_name, sum_then_minus.group(1), depth)
        return total - _resolve(book, sheet_name, sum_then_minus.group(2), depth)

    only_sum = re.fullmatch(r"SUM\((.+?)\)", expression)
    if only_sum:
        return _sum(book, sheet_name, only_sum.group(1), depth)

    difference = re.fullmatch(rf"({_CELL})\s*-\s*({_CELL})", expression)
    if difference:
        return _resolve(book, sheet_name, difference.group(1), depth) - _resolve(
            book, sheet_name, difference.group(2), depth
        )

    if re.fullmatch(_CELL, expression):
        return _resolve(book, sheet_name, expression, depth)

    raise AssertionError(
        f"the evaluator does not understand {expression!r} — the renderer emitted a new formula shape"
    )


def _sum(book, sheet_name: str, argument: str, depth: int) -> float:  # type: ignore[no-untyped-def]
    """Sum a range or a comma-separated list of cells."""
    if ":" in argument:
        start, end = argument.split(":", 1)
        sheet, first_address = _split_cell(start)
        column = re.match(r"[A-Z]{1,3}", first_address).group(0)
        first = int(re.search(r"\d+", first_address).group(0))
        last = int(re.search(r"\d+", _split_cell(end)[1]).group(0))
        return sum(
            evaluate(book, sheet or sheet_name, f"{column}{row}", depth)
            for row in range(first, last + 1)
        )
    return sum(_resolve(book, sheet_name, part, depth) for part in argument.split(","))


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def test_the_workbook_has_the_sheets_the_build_order_asks_for(workbook) -> None:  # type: ignore[no-untyped-def]
    titles = [sheet.title for sheet in workbook.worksheets]
    assert "Summary" in titles
    assert "Business Unit P&L" in titles
    assert "Variance Drivers" in titles
    assert "Incident Impact" in titles
    assert "Lineage" in titles
    assert "Reconciliation" in titles


def test_lineage_and_reconciliation_are_hidden(workbook) -> None:  # type: ignore[no-untyped-def]
    assert workbook["Lineage"].sheet_state == "hidden"
    assert workbook["Reconciliation"].sheet_state == "hidden"
    assert workbook["Business Unit P&L"].sheet_state == "visible"


def test_totals_are_formulas_not_pasted_values(workbook) -> None:  # type: ignore[no-untyped-def]
    """The difference between a workbook and a screenshot of one."""
    sheet = workbook["Business Unit P&L"]
    group_row = next(
        row for row in range(4, sheet.max_row + 1) if sheet.cell(row=row, column=1).value == "Group"
    )
    for column in range(2, 8):
        value = sheet.cell(row=group_row, column=column).value
        assert isinstance(value, str) and value.startswith("=SUM("), (
            f"group total in column {column} is not a formula: {value!r}"
        )

    # Variances too, on every unit row.
    for row in range(4, group_row):
        assert str(sheet.cell(row=row, column=4).value).startswith("="), "revenue variance is pasted"
        assert str(sheet.cell(row=row, column=7).value).startswith("="), "gross profit variance is pasted"


def test_every_formula_evaluates_to_the_fact_it_came_from(rendered: World, workbook) -> None:  # type: ignore[no-untyped-def]
    """The claim that matters. openpyxl does not compute, so this does.

    A renderer could emit syntactically valid nonsense and every other test in
    this file would still pass.
    """
    ir = next(r for r in rendered.artifact_irs if r.tables() and r.tables()[0].key == "summary")
    checked = 0

    for section in ir.sections:
        table = section.table
        if table is None or table.key in ("lineage", "reconciliation", "drivers", "incident_impact"):
            continue
        sheet_name = table.title[:31]
        for row_index, row in enumerate(table.rows):
            for column_index, column in enumerate(table.columns):
                cell = row.cells.get(column.key)
                if cell is None or cell.formula is None or not isinstance(cell.value, (int, float)):
                    continue
                address = f"{chr(ord('B') + column_index)}{4 + row_index}"
                computed = evaluate(workbook, sheet_name, address)
                expected = cell.value / 100 if column.number_format == "0.00%" else cell.value
                assert abs(computed - expected) < 0.01, (
                    f"{sheet_name}!{address} ({row.label}/{column.key}): "
                    f"formula gives {computed}, fact says {expected}"
                )
                checked += 1

    assert checked >= 20, f"only checked {checked} formulas — the sweep is not covering the sheet"


def test_the_reconciliation_sheet_evaluates_to_zero(workbook) -> None:  # type: ignore[no-untyped-def]
    """The workbook checks itself against the ledger when opened."""
    sheet = workbook["Reconciliation"]
    # Filter on the data column, not the label column — the table's trailing note
    # also sits in column 1.
    rows = [r for r in range(4, sheet.max_row + 1) if sheet.cell(row=r, column=2).value is not None]
    assert rows, "no reconciliation checks were written"

    for row in rows:
        summed = sheet.cell(row=row, column=2).value
        stated = sheet.cell(row=row, column=3).value
        difference = sheet.cell(row=row, column=4).value

        assert isinstance(summed, str) and summed.startswith("=SUM("), "the sum must be computed"
        assert isinstance(stated, (int, float)), "the stated total must be a ledger literal"
        assert isinstance(difference, str) and difference.startswith("="), "the check must be computed"
        assert abs(evaluate(workbook, "Reconciliation", f"D{row}")) < 0.01, (
            f"reconciliation row {row} does not net to zero"
        )


def test_the_reconciliation_check_compares_against_the_ledger_not_itself(
    rendered: World, workbook
) -> None:  # type: ignore[no-untyped-def]
    """Otherwise the check is tautological and proves nothing.

    Subtracting the P&L's group cell — itself ``=SUM(units)`` — from a sum of the
    same units can never be non-zero. The comparison has to be against the value
    the fact ledger states, and this test fails if that regresses: corrupt the
    stated literal, and the difference must move.
    """
    sheet = workbook["Reconciliation"]
    assert "Business Unit P&L" not in str(sheet["D4"].value), (
        "the difference must reference the stated literal, not the P&L group cell"
    )

    ir = next(r for r in rendered.artifact_irs if any(t.key == "reconciliation" for t in r.tables()))
    table = next(t for t in ir.tables() if t.key == "reconciliation")
    assert all(row.cells["stated"].fact_id for row in table.rows), (
        "each stated total must be traceable to a fact"
    )

    original = sheet["C4"].value
    sheet["C4"] = original + 1_000
    assert abs(evaluate(workbook, "Reconciliation", "D4") + 1_000) < 0.01, (
        "a wrong stated total must surface as a non-zero difference"
    )
    sheet["C4"] = original


def test_named_ranges_point_at_the_group_totals(rendered: World, workbook) -> None:  # type: ignore[no-untyped-def]
    assert "GroupRevenueActual" in workbook.defined_names
    destination = list(workbook.defined_names["GroupRevenueActual"].destinations)
    sheet_name, address = destination[0]
    assert sheet_name == "Business Unit P&L"

    expected = rendered.facts.where(
        kind="financial.revenue.actual", subject=rendered.company.id
    ).one()
    assert abs(evaluate(workbook, sheet_name, address) - expected.value.amount) < 0.01


def test_number_formats_are_set(workbook) -> None:  # type: ignore[no-untyped-def]
    sheet = workbook["Business Unit P&L"]
    assert "#,##0" in sheet.cell(row=4, column=2).number_format
    assert sheet.cell(row=4, column=8).number_format.endswith("%")


def test_a_percentage_is_stored_as_a_fraction(rendered: World, workbook) -> None:  # type: ignore[no-untyped-def]
    """Excel's percent format multiplies by 100, so 24.94% must be stored as 0.2494."""
    sheet = workbook["Business Unit P&L"]
    computed = evaluate(workbook, "Business Unit P&L", "H4")
    assert 0.0 < computed < 1.0, f"margin stored as {computed}, which would display as a percent of a percent"


def test_the_workbook_declares_itself_synthetic(workbook) -> None:  # type: ignore[no-untyped-def]
    assert "Worldloom" in (workbook.properties.creator or "")
    assert "synthetic" in (workbook.properties.keywords or "").lower()
    assert "Not a real company" in (workbook.properties.description or "")


# ---------------------------------------------------------------------------
# Markdown, and agreement between formats
# ---------------------------------------------------------------------------


def test_markdown_and_xlsx_report_the_same_numbers(rendered: World, workbook) -> None:  # type: ignore[no-untyped-def]
    """Two projections of one IR cannot disagree. This is the property, tested."""
    ir = next(r for r in rendered.artifact_irs if any(t.key == "pnl" for t in r.tables()))
    body = markdown.render(ir).decode("utf-8")

    group = next(row for row in next(t for t in ir.tables() if t.key == "pnl").rows if row.emphasis)
    revenue = group.cells["revenue_actual"].value
    assert f"{revenue:,.0f}" in body

    assert abs(evaluate(workbook, "Business Unit P&L", "C7") - revenue) < 0.01


def test_a_section_awaiting_prose_says_so_rather_than_inventing_it(rendered: World) -> None:
    memo = next(
        ir for ir in rendered.artifact_irs
        if rendered.artifact_intents.by_id(ir.intent_id).artifact_type == "cfo_variance_memo"
    )
    body = markdown.render(memo).decode("utf-8")
    assert "Awaiting narrative" in body
    assert "## Position" in body, "the outline's headings should already be final"
    assert memo.metadata["awaiting_prose"] == "true"


def test_narrative_outlines_bind_their_facts_before_prose_exists(rendered: World) -> None:
    for ir in rendered.artifact_irs:
        intent = rendered.artifact_intents.by_id(ir.intent_id)
        if intent.artifact_type == "finance_workbook":
            continue
        assert ir.fact_ids(), f"{ir.id} has no facts bound"
        assert set(ir.fact_ids()) <= set(intent.required_fact_ids)


def test_every_rendered_artifact_carries_the_synthetic_notice(rendered: World) -> None:
    for item in rendered._rendered:
        if item.media_type == "text/markdown":
            assert "Not a real company" in item.text


# ---------------------------------------------------------------------------
# Portable bundles
# ---------------------------------------------------------------------------


def _bundle(rendered: World, path: str) -> list[dict]:
    item = next(r for r in rendered._rendered if r.path == path)
    if path.endswith(".jsonl"):
        return [json.loads(line) for line in item.text.splitlines() if line.strip()]
    return [json.loads(item.text)]


def test_jira_separates_the_control_fix_from_the_detection_fix(rendered: World) -> None:
    """A single undifferentiated ticket cannot pose the question of which fix works."""
    issues = _bundle(rendered, "jira/issues.jsonl")
    assert len(issues) == 2
    assert {issue["addresses"] for issue in issues} == {"detection_gap", "control_failure"}

    control = next(i for i in issues if i["addresses"] == "control_failure")
    assert control["priority"] == "Highest"
    assert control["worldloom_fact_ids"], "the control fix must cite the classification facts"


def test_jira_links_back_to_the_incident(rendered: World) -> None:
    links = _bundle(rendered, "jira/links.jsonl")
    assert links
    assert all(link["system"] == "servicenow" for link in links)
    assert all(link["to"].startswith("INC") for link in links)


def test_servicenow_work_notes_preserve_the_ruled_out_hypothesis(rendered: World) -> None:
    """A single root-cause field cannot express what was believed when."""
    incident = _bundle(rendered, "servicenow/incident.json")[0]
    notes = incident["work_notes"]

    superseded = [note for note in notes if note["superseded"]]
    assert superseded, "the hypothesis that was ruled out must survive in the record"
    assert any(note["authority"] == "initial_hypothesis" for note in notes)
    assert any(note["authority"] == "confirmed" for note in notes)

    moments = [note["at"] for note in notes]
    assert moments == sorted(moments), "work notes must be in time order"


def test_servicenow_ships_the_cmdb_the_incident_points_at(rendered: World) -> None:
    incident = _bundle(rendered, "servicenow/incident.json")[0]
    cmdb = _bundle(rendered, "servicenow/cmdb_ci.jsonl")
    assert incident["cmdb_ci"] in {item["sys_id"] for item in cmdb}

    relations = _bundle(rendered, "servicenow/cmdb_rel_ci.jsonl")
    assert relations, "service dependencies should be exported"


def test_confluence_labels_a_stale_page_as_stale(rendered: World) -> None:
    pages = _bundle(rendered, "confluence/pages.jsonl")
    stale = [page for page in pages if page["stale"]]
    assert stale, "the triage page cites a superseded fact and should be marked"

    comments = _bundle(rendered, "confluence/comments.jsonl")
    assert {c["page"] for c in comments} <= {p["id"] for p in pages}


def test_bundles_are_plain_text_a_consumer_can_read_without_us(rendered: World) -> None:
    for item in rendered._rendered:
        if not item.path.startswith(("jira/", "confluence/", "servicenow/")):
            continue
        assert item.media_type in ("application/json", "application/jsonl")
        if item.path.endswith(".jsonl"):
            for line in item.text.splitlines():
                if line.strip():
                    json.loads(line)
        else:
            json.loads(item.text)


# ---------------------------------------------------------------------------
# Manifest, export, and validation
# ---------------------------------------------------------------------------


def test_rendering_produces_a_manifest_entry_per_artifact(rendered: World) -> None:
    assert len(rendered.artifacts) == len(rendered.artifact_intents)
    assert len(set(rendered.artifacts.ids())) == len(rendered.artifacts)


def test_no_artifact_is_dated_before_the_facts_it_cites(rendered: World) -> None:
    """Derived rather than chosen, so this cannot drift."""
    for artifact in rendered.artifacts:
        for fact_id in artifact.supporting_fact_ids:
            assert rendered.facts.by_id(fact_id).valid_from <= artifact.created_at


def test_authority_differs_by_artifact_type(rendered: World) -> None:
    by_type = {a.artifact_type: a.authority.value for a in rendered.artifacts}
    assert by_type["finance_workbook"] == "system_of_record"
    assert by_type["cfo_variance_memo"] == "approved_report"
    assert by_type["confluence_page"] == "unofficial_note"


def test_an_unknown_audience_does_not_publish_to_all_staff(rendered: World) -> None:
    """Falling open would leak; falling closed only annoys."""
    policies = {p.id: p.label for p in rendered.access_policies}
    for artifact in rendered.artifacts:
        if artifact.audience == "all_staff":
            continue
        assert policies[artifact.access_policy_id] != "All staff", artifact.id


def test_a_rendered_world_validates(rendered: World) -> None:
    report = rendered.validate()
    assert report.ok, "\n".join(str(v) for v in report.violations)


def test_exported_files_exist_and_the_reloaded_corpus_validates(rendered: World, tmp_path) -> None:
    """The file check is inert without a manifest, so this is where it bites."""
    destination = rendered.export(tmp_path / "out")

    for artifact in rendered.artifacts:
        assert (destination / artifact.path).is_file(), artifact.path

    reloaded = World.load(destination)
    assert len(reloaded.artifacts) == len(rendered.artifacts)
    assert len(reloaded.artifact_irs) == len(rendered.artifact_irs)
    report = reloaded.validate()
    assert report.ok, "\n".join(str(v) for v in report.violations)


def test_rendering_does_not_mutate_the_source_world() -> None:
    world = RetailWorld(seed=8128).build().run(MonthEndClose(period=PERIOD, include_operational_incident=True))
    world.render("markdown")
    assert len(world.artifacts) == 0
    assert len(world.artifact_irs) == 0


def test_render_is_deterministic() -> None:
    def build():  # type: ignore[no-untyped-def]
        return (
            RetailWorld(seed=8128).build()
            .run(MonthEndClose(period=PERIOD, include_operational_incident=True))
            .render("markdown", "jira")
        )

    first = {item.path: item.payload for item in build()._rendered}
    second = {item.path: item.payload for item in build()._rendered}
    assert first == second


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_formats_are_registered_as_plugins() -> None:
    assert set(available()) == {"markdown", "xlsx", "jira", "confluence", "servicenow"}


def test_an_unknown_format_says_what_is_available(rendered: World) -> None:
    with pytest.raises(RenderError, match="unknown format"):
        rendered.render("powerpoint")


def test_rendering_nothing_is_an_error() -> None:
    world = RetailWorld(seed=8128).build()
    with pytest.raises(ValueError, match="run a scenario first"):
        world.render("markdown")

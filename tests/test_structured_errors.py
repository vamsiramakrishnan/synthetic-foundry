"""Mechanical spreadsheet errors: planned in the ledger, real on the page.

The claim under test runs end to end. The messiness pass mints a
``workbook_copy`` intent and two mechanical ``IntentionalError`` records
(`hardcoded_value`, `short_range`), seeded from the world alone;
``compiler.mechanical`` makes each record true of one cell of the compiled
copy; ``render.xlsx`` — untouched — emits the typed-in literal and the
truncated range; and ``validate.intentional`` refuses any label the compiled
sheet does not substantiate. The default build mints none of it, which is the
half of the contract that keeps every existing corpus byte-identical.
"""

from __future__ import annotations

import io
import re
from dataclasses import replace

import openpyxl
import pytest

from worldloom import MonthEndClose, RetailWorld, World
from worldloom import messiness as messiness_module
from worldloom.compiler import mechanical
from worldloom.generators import distractors
from worldloom.models import ErrorType, FormulaKind

PERIOD = "2026-03"
SEED = 8128

#: The explicit configuration that reaches the new dimension. A budget mapping
#: rather than a named profile, deliberately: every named profile keeps
#: ``mechanical`` at zero so no corpus gains a wrong workbook without asking.
BUDGET = {"mechanical": 2}


def _base() -> World:
    return RetailWorld(seed=SEED).build().run(
        MonthEndClose(period=PERIOD, include_operational_incident=True)
    )


def _mechanical_errors(world: World):  # type: ignore[no-untyped-def]
    return [
        error for error in world.intentional_errors
        if error.error_type in mechanical.MECHANICAL_KINDS
    ]


@pytest.fixture(scope="module")
def messy() -> World:
    return messiness_module.apply(_base(), BUDGET)


@pytest.fixture(scope="module")
def rendered(messy: World) -> World:
    return messy.render("xlsx", "markdown")


def _workbooks(rendered_world: World):  # type: ignore[no-untyped-def]
    """``(copy sheet, source sheet, copy IR, errors)`` for the corrupted pair."""
    copy_intent = next(
        intent for intent in rendered_world.artifact_intents
        if intent.artifact_type == mechanical.WORKBOOK_COPY
    )
    source_intent = next(
        intent for intent in rendered_world.artifact_intents
        if intent.artifact_type == "finance_workbook"
    )
    ir_of = {ir.intent_id: ir for ir in rendered_world.artifact_irs}
    payloads = {r.artifact_id: r.payload for r in rendered_world._rendered if r.path.endswith(".xlsx")}
    copy = openpyxl.load_workbook(io.BytesIO(payloads[ir_of[copy_intent.id].id]))
    source = openpyxl.load_workbook(io.BytesIO(payloads[ir_of[source_intent.id].id]))
    return copy, source, ir_of[copy_intent.id], _mechanical_errors(rendered_world)


# ---------------------------------------------------------------------------
# Plan time
# ---------------------------------------------------------------------------


def test_minting_is_deterministic() -> None:
    """Two builds of one seed plan identical errors, ids and all.

    Byte-for-byte at the model level, which is what the corpus's jsonl is
    serialised from — the property CI's ledger-regeneration diff enforces on
    whole corpora, asserted here at the source.
    """
    first = messiness_module.apply(_base(), BUDGET)
    second = messiness_module.apply(_base(), BUDGET)
    a, b = _mechanical_errors(first), _mechanical_errors(second)
    assert a and [e.model_dump() for e in a] == [e.model_dump() for e in b]
    copies_a = [i for i in first.artifact_intents if i.artifact_type == mechanical.WORKBOOK_COPY]
    copies_b = [i for i in second.artifact_intents if i.artifact_type == mechanical.WORKBOOK_COPY]
    assert copies_a and [i.model_dump() for i in copies_a] == [i.model_dump() for i in copies_b]


def test_both_kinds_are_planned_and_each_names_its_disagreement(messy: World) -> None:
    errors = _mechanical_errors(messy)
    assert {error.error_type for error in errors} == set(mechanical.MECHANICAL_KINDS)
    facts = {fact.id: fact for fact in messy.facts}
    for error in errors:
        observed = mechanical.parsed(error.observed_value)
        assert observed is not None, "mechanical observed_value must be a bare reading"
        canonical = facts[error.canonical_fact_id]
        assert canonical.value is not None
        # The record names a real disagreement, not a restatement.
        assert abs(observed - canonical.value.amount) > 0.01
        # And the copy, not the system of record, carries it.
        labelled = messy.artifact_intents.by_id(error.artifact_id)
        assert labelled.artifact_type == mechanical.WORKBOOK_COPY
        assert labelled.derived_from, "the copy must point at the workbook it copies"


def test_measuring_the_ceiling_mints_nothing() -> None:
    """`messiness_ceilings` is a reading, and a reading must not move the world.

    `_orphanable`'s docstring records the defect class: a measurement that
    advances the id sequence makes a corpus replay differently from its own
    build. So the errors planned after a ceilings call must be identical to
    those planned without one.
    """
    counted = _base()
    ceilings = distractors.messiness_ceilings(counted)
    assert ceilings["mechanical"] >= 2
    after_measuring = messiness_module.apply(counted, BUDGET)
    fresh = messiness_module.apply(_base(), BUDGET)
    assert [e.model_dump() for e in _mechanical_errors(after_measuring)] == [
        e.model_dump() for e in _mechanical_errors(fresh)
    ]


# ---------------------------------------------------------------------------
# The page
# ---------------------------------------------------------------------------


def test_the_hardcoded_cell_is_a_literal_where_siblings_derive(rendered: World) -> None:
    """The paste-over is real: a number sits where every other row has ``=``."""
    copy, source, copy_ir, errors = _workbooks(rendered)
    error = next(e for e in errors if e.error_type is ErrorType.HARDCODED_VALUE)
    observed = mechanical.parsed(error.observed_value)

    pnl = next(table for table in copy_ir.tables() if table.key == "pnl")
    row_index = next(
        index for index, row in enumerate(pnl.rows)
        if any(cell.fact_id == error.canonical_fact_id and cell.formula is None
               for cell in row.cells.values())
    )
    column_index = next(
        index for index, column in enumerate(pnl.columns) if column.key == "revenue_variance"
    )
    address_row = 4 + row_index
    address_column = 2 + column_index

    corrupted = copy["Business Unit P&L"].cell(row=address_row, column=address_column).value
    original = source["Business Unit P&L"].cell(row=address_row, column=address_column).value
    assert isinstance(corrupted, (int, float)) and abs(corrupted - observed) < 0.01, (
        f"the copy's cell should be the typed-in {observed}, not {corrupted!r}"
    )
    assert isinstance(original, str) and original.startswith("="), (
        "the system of record must keep its formula — the disagreement is the product"
    )


def test_the_short_range_stops_one_row_early(rendered: World) -> None:
    copy, source, copy_ir, errors = _workbooks(rendered)
    error = next(e for e in errors if e.error_type is ErrorType.SHORT_RANGE)

    pnl = next(table for table in copy_ir.tables() if table.key == "pnl")
    row_index = next(
        index for index, row in enumerate(pnl.rows)
        if any(cell.fact_id == error.canonical_fact_id and cell.formula is FormulaKind.SUM
               for cell in row.cells.values())
    )
    column_index = next(
        index for index, column in enumerate(pnl.columns) if column.key == "revenue_actual"
    )
    address_row, address_column = 4 + row_index, 2 + column_index

    corrupted = copy["Business Unit P&L"].cell(row=address_row, column=address_column).value
    original = source["Business Unit P&L"].cell(row=address_row, column=address_column).value
    short = re.fullmatch(r"=SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)", str(corrupted))
    full = re.fullmatch(r"=SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)", str(original))
    assert short and full, f"both totals should be ranges: {corrupted!r} vs {original!r}"
    assert int(short.group(2)) == int(full.group(2)) - 1, (
        f"the copy's range should stop exactly one row early: {corrupted!r} vs {original!r}"
    )
    assert short.group(1) == full.group(1), "the range should start where the real one starts"


def test_the_copy_agrees_with_itself_across_formats(rendered: World) -> None:
    """The corrupted cell's IR literal is the wrong number the formula computes.

    Markdown emits the literal where XLSX emits the formula, so a literal left
    canonical would make one document say two things — the engine-wide rule the
    corruption must not break. One cell disagreeing with the *ledger* is the
    product; one document disagreeing with itself would be a defect.
    """
    _copy, _source, copy_ir, errors = _workbooks(rendered)
    for error in errors:
        observed = mechanical.parsed(error.observed_value)
        stated = [
            cell.value
            for table in copy_ir.tables()
            for row in table.rows
            for cell in row.cells.values()
            if cell.fact_id == error.canonical_fact_id
            and isinstance(cell.value, (int, float))
            and abs(float(cell.value) - observed) < 0.01
        ]
        assert stated, f"{error.id}: the IR literal must carry the wrong reading too"


# ---------------------------------------------------------------------------
# The validator
# ---------------------------------------------------------------------------


def test_validate_substantiates_the_labels(messy: World) -> None:
    """The compiled corpus passes whole — the planned errors are recognised via
    the ledger, not flagged by any formula or evidence check."""
    report = messy.compile().validate()
    assert report.ok, [str(v) for v in report.violations[:8]]


def test_validate_refuses_a_label_the_sheet_does_not_substantiate(messy: World) -> None:
    """The check has teeth: shift a recorded reading and the label is refused.

    This is the "refuses a label the corpus cannot substantiate" half of the
    contract — without it a mislabelled corruption would be the corpus vouching
    for a disagreement it cannot show.
    """
    compiled = messy.compile()
    target = _mechanical_errors(compiled)[0]
    tampered = tuple(
        error.model_copy(update={
            "observed_value": mechanical.reading(mechanical.parsed(error.observed_value) + 1.0)
        })
        if error.id == target.id else error
        for error in compiled._intentional_errors
    )
    report = replace(compiled, _intentional_errors=tampered).validate()
    assert not report.ok
    assert any(v.code == "mechanical_unsubstantiated" for v in report.violations), (
        "the tampered label should be refused by name"
    )


def test_a_label_claiming_no_disagreement_is_refused(messy: World) -> None:
    """observed == canonical is not an imperfection, whatever the label says."""
    compiled = messy.compile()
    facts = {fact.id: fact for fact in compiled.facts}
    target = _mechanical_errors(compiled)[0]
    canonical = facts[target.canonical_fact_id].value.amount
    tampered = tuple(
        error.model_copy(update={"observed_value": mechanical.reading(canonical)})
        if error.id == target.id else error
        for error in compiled._intentional_errors
    )
    report = replace(compiled, _intentional_errors=tampered).validate()
    assert any(v.code == "mechanical_unsubstantiated" for v in report.violations)


# ---------------------------------------------------------------------------
# The default stays the default
# ---------------------------------------------------------------------------


def test_a_default_build_mints_none_of_it() -> None:
    world = _base()
    assert not _mechanical_errors(world)
    assert not [
        intent for intent in world.artifact_intents
        if intent.artifact_type == mechanical.WORKBOOK_COPY
    ]


def test_every_named_profile_keeps_the_new_dimension_at_zero() -> None:
    """The dimension is opt-in by budget. A named profile that quietly turned
    it on would put a wrong workbook into corpora built long before the kind
    existed — the byte-identity contract `KINDS`' comment states."""
    for name, profile in messiness_module.PROFILES.items():
        assert profile["mechanical"] == 0, name


def test_replays_from_its_own_recipe() -> None:
    """The messiness step records the budget, so the corpus rebuilds itself."""
    from worldloom import recipe as recipe_module

    built = messiness_module.apply(_base(), BUDGET)
    again = recipe_module.rebuild(built._recipe)
    assert [e.model_dump() for e in _mechanical_errors(built)] == [
        e.model_dump() for e in _mechanical_errors(again)
    ]
    assert [i.model_dump() for i in built.artifact_intents] == [
        i.model_dump() for i in again.artifact_intents
    ]

#!/usr/bin/env python3
"""Harvest external taxonomy/distribution sources into Worldloom's lexicon schema.

Raw source files are inputs, not repository assets. This command emits compact
JSONL/prior files plus a license/provenance ledger. Network download is kept in
workflow/shell code so parsing remains deterministic and testable offline.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def lexicon_row(
    id_: str,
    type_: str,
    label: str,
    source: str,
    license_: str,
    **extra: Any,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "id": id_,
        "type": type_,
        "label": label,
        "canonical": None,
        "alt_labels": [],
        "lang": "en",
        "industry": None,
        "region": None,
        "weight": 1.0,
        "source": source,
        "license": license_,
    }
    row.update(extra)
    return row


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
            count += 1
    return count


def harvest_apqc(path: Path, out: Path, *, source: str, license_: str) -> int:
    """Parse an APQC PCF workbook already downloaded under its source terms."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for cells in sheet.iter_rows(values_only=True):
        values = [str(cell or "").strip() for cell in cells]
        if header is None and any("Hierarchy" in value or value in {"PCF ID", "Process ID"} for value in values):
            header = values
            continue
        if header is None or not any(values):
            continue
        data = dict(zip(header, values, strict=False))
        hierarchy_id = data.get("Hierarchy ID") or data.get("PCF ID") or data.get("Process ID") or ""
        label = data.get("Name") or data.get("Process Element") or data.get("Process Name") or ""
        if not hierarchy_id or not label:
            continue
        level = hierarchy_id.count(".") + 1
        rows.append(
            lexicon_row(
                f"{source}:{hierarchy_id}",
                "process" if level <= 3 else "activity",
                label,
                source,
                license_,
                canonical=f"apqc:{hierarchy_id}",
                weight=1.0 / level,
                hierarchy_id=hierarchy_id,
                hierarchy_level=level,
            )
        )
    return write_jsonl(out / f"{source}.jsonl", rows)


def harvest_esco_csv(path: Path, out: Path, *, source: str = "esco") -> int:
    """Parse an ESCO occupation CSV and retain its ISCO crosswalk."""
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for record in csv.DictReader(handle):
            uri = record.get("conceptUri") or record.get("Concept URI") or ""
            label = record.get("preferredLabel") or record.get("Concept PT") or ""
            isco = record.get("iscoGroup") or record.get("ISCO code") or record.get("Parent ISCO code") or ""
            if not uri or not label:
                continue
            concept_id = uri.rsplit("/", 1)[-1]
            alt_raw = record.get("altLabels") or record.get("Alt Labels") or ""
            alt_labels = [
                {"lang": "en", "label": value.strip()}
                for value in re.split(r"[\n|]", alt_raw)
                if value.strip()
            ]
            rows.append(
                lexicon_row(
                    f"esco:{concept_id}",
                    "title",
                    label,
                    source,
                    "CC BY 4.0 (European Commission ESCO)",
                    canonical=f"isco:{isco}" if isco else f"esco:{concept_id}",
                    alt_labels=alt_labels,
                    description=(record.get("description") or record.get("Definition") or "")[:500],
                    isco=isco or None,
                )
            )
    return write_jsonl(out / "esco.jsonl", rows)


def _walk_json(value: Any) -> Iterable[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from _walk_json(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk_json(child)


def harvest_jira_jsonl(path: Path, out: Path) -> int:
    """Derive field fill rates and issue-type mix from an anonymised Jira JSONL export."""
    fields: Counter[str] = Counter()
    issue_types: Counter[str] = Counter()
    names: dict[str, str] = {}
    issues = 0
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                issue = json.loads(line)
            except ValueError:
                continue
            fs = issue.get("fields") or {}
            if not isinstance(fs, dict):
                continue
            issues += 1
            issue_type = fs.get("issuetype") or {}
            if isinstance(issue_type, dict):
                issue_types[str(issue_type.get("name") or "unknown")] += 1
            for key, value in fs.items():
                if key.startswith("customfield_") and value not in (None, [], "", {}):
                    fields[key] += 1
            for key, value in (issue.get("names") or {}).items():
                names[str(key)] = str(value)

    denominator = max(issues, 1)
    rows = [
        lexicon_row(
            f"jira:{field}",
            "field",
            names.get(field, field),
            "public-jira-dataset-anonymised",
            "See Zenodo record 15719919 and dataset terms",
            canonical=None,
            weight=count / denominator,
            fill_rate=count / denominator,
        )
        for field, count in fields.most_common()
    ]
    rows.extend(
        lexicon_row(
            f"jira:issuetype:{slug(name)}",
            "state",
            name,
            "public-jira-dataset-anonymised",
            "See Zenodo record 15719919 and dataset terms",
            canonical="issue.type",
            weight=count / denominator,
        )
        for name, count in issue_types.most_common()
    )
    write_jsonl(out / "jira-fields.jsonl", rows)
    stats = {
        "source": "public-jira-dataset-anonymised",
        "issues": issues,
        "custom_fields_observed": len(fields),
        "issue_types": dict(issue_types.most_common()),
    }
    (out / "jira-stats.json").write_text(json.dumps(stats, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return issues


def harvest_xes(path: Path, out: Path, *, source: str) -> int:
    """Stream an XES/XES.GZ log into activity, transition, timing and rework priors."""
    opener = gzip.open if path.suffix == ".gz" else open
    activities: Counter[str] = Counter()
    transitions: Counter[str] = Counter()
    rework_cases = 0
    cases = 0
    durations_seconds: list[float] = []

    with opener(path, "rb") as handle:
        tree = ET.iterparse(handle, events=("end",))
        for _, element in tree:
            if not element.tag.endswith("trace"):
                continue
            cases += 1
            sequence: list[tuple[str, str | None]] = []
            for event in element:
                if not event.tag.endswith("event"):
                    continue
                name: str | None = None
                timestamp: str | None = None
                for attr in event:
                    key = attr.attrib.get("key")
                    if key == "concept:name":
                        name = attr.attrib.get("value")
                    elif key == "time:timestamp":
                        timestamp = attr.attrib.get("value")
                if name:
                    sequence.append((name, timestamp))
                    activities[name] += 1
            if len({name for name, _ in sequence}) < len(sequence):
                rework_cases += 1
            for (left, left_ts), (right, right_ts) in zip(sequence, sequence[1:], strict=False):
                transitions[f"{left} -> {right}"] += 1
                if left_ts and right_ts:
                    try:
                        from datetime import datetime

                        delta = (datetime.fromisoformat(right_ts.replace("Z", "+00:00")) - datetime.fromisoformat(left_ts.replace("Z", "+00:00"))).total_seconds()
                        if delta >= 0:
                            durations_seconds.append(delta)
                    except ValueError:
                        pass
            element.clear()

    event_count = sum(activities.values())
    activity_rows = [
        lexicon_row(
            f"{source}:activity:{slug(name)}",
            "activity",
            name,
            source,
            "Per-log BPI/4TU license; preserve source DOI",
            canonical=f"{source}:activity:{slug(name)}",
            weight=count / max(event_count, 1),
        )
        for name, count in activities.most_common()
    ]
    write_jsonl(out / f"{source}-activities.jsonl", activity_rows)
    ordered = sorted(durations_seconds)
    quantile = lambda q: ordered[min(int(q * (len(ordered) - 1)), len(ordered) - 1)] if ordered else None
    stats = {
        "source": source,
        "cases": cases,
        "events": event_count,
        "rework_case_rate": rework_cases / max(cases, 1),
        "state_transitions": dict(transitions.most_common(100)),
        "inter_activity_seconds": {
            "p50": quantile(0.50),
            "p90": quantile(0.90),
            "p99": quantile(0.99),
        },
    }
    (out / f"{source}-priors.json").write_text(json.dumps(stats, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return event_count


def write_provenance(out: Path, source: str, path: Path, *, license_: str) -> None:
    ledger = out / "licenses.jsonl"
    row = {
        "source": source,
        "input": path.name,
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
        "license": license_,
    }
    with ledger.open("a", encoding="utf-8", newline="\n") as handle:
        handle.write(json.dumps(row, sort_keys=True) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, required=True)
    sub = parser.add_subparsers(dest="command", required=True)

    apqc = sub.add_parser("apqc")
    apqc.add_argument("path", type=Path)
    apqc.add_argument("--source", required=True)
    apqc.add_argument("--license", required=True)

    esco = sub.add_parser("esco-csv")
    esco.add_argument("path", type=Path)

    jira = sub.add_parser("jira-jsonl")
    jira.add_argument("path", type=Path)

    xes = sub.add_parser("xes")
    xes.add_argument("path", type=Path)
    xes.add_argument("--source", required=True)

    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)
    if args.command == "apqc":
        count = harvest_apqc(args.path, args.out, source=args.source, license_=args.license)
        write_provenance(args.out, args.source, args.path, license_=args.license)
    elif args.command == "esco-csv":
        count = harvest_esco_csv(args.path, args.out)
        write_provenance(args.out, "esco", args.path, license_="CC BY 4.0")
    elif args.command == "jira-jsonl":
        count = harvest_jira_jsonl(args.path, args.out)
        write_provenance(args.out, "public-jira-dataset-anonymised", args.path, license_="See Zenodo 15719919")
    else:
        count = harvest_xes(args.path, args.out, source=args.source)
        write_provenance(args.out, args.source, args.path, license_="Per-log BPI/4TU license")
    print(json.dumps({"command": args.command, "rows_or_events": count, "out": str(args.out)}, sort_keys=True))


if __name__ == "__main__":
    main()
